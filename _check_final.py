"""修复后的全面检查"""
import subprocess
from openpyxl import load_workbook
from amazon_excel_processor.merger import merge_files, _group_base_name, VARIANT_GROUP_SIZE, MAIN_GROUP_SIZE, _col_letter, COL_PARENT_SKU, COL_SELLER_SKU
from amazon_excel_processor.excel_io import group_rows

print("=" * 70)
print("检查 0: 全量测试")
print("=" * 70)
r = subprocess.run(["poetry", "run", "pytest", "tests/", "-q"], capture_output=True, text=True)
print(r.stdout.strip().split("\n")[-1])

print("\n生成 2 种模式的输出...")
out_new = merge_files("lh725测试HM普.xlsm", "lh725测试HM木.xlsm", "lh725测试HM金_split.xlsm",
                      sku_prefix="HM725", mode="new", output_path="/tmp/check_new.xlsm")
out_old = merge_files("lh725测试HM普.xlsm", "lh725测试HM木.xlsm", "lh725测试HM金_split.xlsm",
                      sku_prefix="HM725", mode="old_variant", output_path="/tmp/check_old.xlsm")
print(f"新品上架: {out_new}")
print(f"老品补充: {out_old}")

ws_new = load_workbook(out_new, keep_vba=True)["Template"]
ws_old = load_workbook(out_old, keep_vba=True)["Template"]
ws_fin = load_workbook("lh725测试HM最终.xlsm", keep_vba=True)["Template"]
ws_main = load_workbook("lh725测试HM普.xlsm", keep_vba=True)["Template"]
ws_wood = load_workbook("lh725测试HM木.xlsm", keep_vba=True)["Template"]
ws_gold = load_workbook("lh725测试HM金_split.xlsm", keep_vba=True)["Template"]

def _empty(v):
    if v is None: return ""
    if isinstance(v, str): return v.strip()
    return v

# 检查 1: 新品上架 vs FINAL (关键列 + Image URL, 排除 None/"" 差异)
print("\n" + "=" * 70)
print("检查 1: 新品上架 vs FINAL (关键 14 列 + 9 Image URL 列)")
print("=" * 70)
cols = {
    'sku': 2, 'name': 9, 'price': 13, 'list': 145, 'relType': 24,
    'varThm': 26, 'parSKU': 27, 'par': 30, 'color': 38, 'size': 41,
    'sizeMap': 55, 'len': 62, 'wid': 63, 'wt': 69,
    'main_img': 14, 'img1': 15, 'img2': 16, 'img3': 17, 'img4': 18,
    'img5': 19, 'img6': 20, 'img7': 21, 'swatch': 23,
}
mism = 0; total = 0
for r in range(4, 88):
    for cn, c in cols.items():
        ov = _empty(ws_new.cell(row=r, column=c).value)
        fv = _empty(ws_fin.cell(row=r, column=c).value)
        if ov == "" and fv == "": continue
        total += 1
        if ov != fv:
            mism += 1
            if mism <= 5: print(f"  ✗ r{r} {cn}: OUT={str(ov)[:25]} vs FIN={str(fv)[:25]}")
print(f"  总 {total}, 不一致 {mism}, 一致 {total-mism} ({100*(total-mism)/total:.1f}%) {'✓' if mism==0 else '✗'}")

# 检查 2: 新品上架 Parent SKU 公式链
print("\n" + "=" * 70)
print("检查 2: 新品上架 Parent SKU 公式链 (第 1 画)")
print("=" * 70)
sl = _col_letter(COL_SELLER_SKU); pl = _col_letter(COL_PARENT_SKU)
ok = True
for r in range(4, 25):
    ps = ws_new.cell(row=r, column=27).value
    exp = None if r == 4 else (f"={sl}4" if r == 5 else f"={pl}{r-1}")
    if ps != exp: ok = False; print(f"  ✗ r{r}: {ps} vs {exp}")
print(f"  {'✓' if ok else '✗'} 公式链正确 (r5=B4, r6-r24=AA{{prev}})")

# 检查 3: 老品补充 普文件原 11 行完全不动 (全 179 列)
print("\n" + "=" * 70)
print("检查 3: 老品补充 普文件原 11 行完全不动 (全 179 列 × 4 画)")
print("=" * 70)
real_diff = 0
for gi in range(4):
    for i in range(11):
        r_out = 4 + gi * 21 + i
        r_main = 4 + gi * 11 + i
        for c in range(1, 180):
            ov = _empty(ws_old.cell(row=r_out, column=c).value)
            mv = _empty(ws_main.cell(row=r_main, column=c).value)
            if ov == "" and mv == "": continue
            if ov != mv:
                real_diff += 1
                if real_diff <= 3:
                    h = ws_main.cell(row=2, column=c).value or f"col{c}"
                    print(f"  ✗ 画{gi+1} r{r_out}[{h}]: OUT={str(ov)[:20]} vs MAIN={str(mv)[:20]}")
print(f"  真正不一致: {real_diff} {'✓ 普文件原数据完全保留' if real_diff == 0 else '✗'}")

# 检查 4: 老品补充 Wood/Gold 行 SKU + parSKU
print("\n" + "=" * 70)
print("检查 4: 老品补充 Wood/Gold SKU + parSKU (第 1 画 r15-r24)")
print("=" * 70)
ok = True
for i in range(10):
    r = 15 + i
    sku = ws_old.cell(row=r, column=2).value
    ps = ws_old.cell(row=r, column=27).value
    exp_sku = f"HM725-{i+1}"
    exp_ps = f"={pl}{r-1}"
    if sku != exp_sku: ok = False; print(f"  ✗ r{r} sku={sku} vs {exp_sku}")
    if ps != exp_ps: ok = False; print(f"  ✗ r{r} parSKU={ps} vs {exp_ps}")
print(f"  {'✓' if ok else '✗'} Wood/Gold SKU=HM725-1~10, parSKU=AA{{prev}} 链式")

# 检查 5: 老品补充 Wood/Gold 行 Color + Price
print("\n" + "=" * 70)
print("检查 5: 老品补充 Wood/Gold Color + Price (第 1 画)")
print("=" * 70)
ok = True
for i in range(5):
    r = 15 + i
    color = ws_old.cell(row=r, column=38).value
    price = ws_old.cell(row=r, column=13).value
    if color != "Vintage Wood Grain Frame-style": ok = False; print(f"  ✗ r{r} color={color}")
    if price != [26.9, 39.9, 59.9, 99.9, 129.9][i]: ok = False; print(f"  ✗ r{r} price={price}")
for i in range(5):
    r = 20 + i
    color = ws_old.cell(row=r, column=38).value
    price = ws_old.cell(row=r, column=13).value
    if color != "Vintage Ornate Gold Frame-style": ok = False; print(f"  ✗ r{r} color={color}")
    if price != [26.9, 39.9, 59.9, 99.9, 129.9][i]: ok = False; print(f"  ✗ r{r} price={price}")
print(f"  {'✓' if ok else '✗'} Wood/Gold Color + Price 正确")

# 检查 6: 老品补充 Wood/Gold List Price = Your Price
print("\n" + "=" * 70)
print("检查 6: 老品补充 Wood/Gold List Price = Your Price")
print("=" * 70)
ok = True
for r in range(15, 25):
    yp = ws_old.cell(row=r, column=13).value
    lp = ws_old.cell(row=r, column=145).value
    if yp != lp: ok = False; print(f"  ✗ r{r}: YP={yp} vs LP={lp}")
print(f"  {'✓' if ok else '✗'} List Price = Your Price")

# 检查 7: 老品补充 Wood/Gold Parentage/RelType
print("\n" + "=" * 70)
print("检查 7: 老品补充 Wood/Gold Parentage/RelType (全 4 画)")
print("=" * 70)
ok = True
for gi in range(4):
    for i in range(10):
        r = 4 + gi * 21 + 11 + i
        par = ws_old.cell(row=r, column=30).value
        rel = ws_old.cell(row=r, column=24).value
        if par != "Child": ok = False; print(f"  ✗ r{r} par={par}")
        if rel != "Variation": ok = False; print(f"  ✗ r{r} rel={rel}")
print(f"  {'✓' if ok else '✗'} Wood/Gold 全部 Child/Variation")

# 检查 8: 老品补充 Image URL 来源 (按正确配对)
print("\n" + "=" * 70)
print("检查 8: 老品补充 Image URL 来源 (按 base name 配对)")
print("=" * 70)
main_groups = group_rows(ws_main, group_size=MAIN_GROUP_SIZE)
wood_groups = group_rows(ws_wood, group_size=VARIANT_GROUP_SIZE)
gold_groups = group_rows(ws_gold, group_size=VARIANT_GROUP_SIZE)
wood_by_name = {_group_base_name(ws_wood, g): g for g in wood_groups}
gold_by_name = {_group_base_name(ws_gold, g): g for g in gold_groups}
ok = True
for gi, mg in enumerate(main_groups):
    name = _group_base_name(ws_main, mg)
    wg = wood_by_name.get(name); gg = gold_by_name.get(name)
    out_start = 4 + gi * 21
    for i in range(5):
        if ws_old.cell(row=out_start+11+i, column=14).value != ws_wood.cell(row=wg[1+i], column=14).value:
            ok = False; print(f"  ✗ 画{gi+1} r{out_start+11+i} Wood img")
        if ws_old.cell(row=out_start+16+i, column=14).value != ws_gold.cell(row=gg[1+i], column=14).value:
            ok = False; print(f"  ✗ 画{gi+1} r{out_start+16+i} Gold img")
print(f"  {'✓' if ok else '✗'} Image URL 来源正确 (Wood←木文件, Gold←金文件)")

# 检查 9: 老品补充 普文件原行 parSKU 保留
print("\n" + "=" * 70)
print("检查 9: 老品补充 普文件原 11 行 parSKU 保留 (全 4 画)")
print("=" * 70)
ok = True
for gi in range(4):
    for i in range(11):
        r_out = 4 + gi * 21 + i
        r_main = 4 + gi * 11 + i
        op = _empty(ws_old.cell(row=r_out, column=27).value)
        mp = _empty(ws_main.cell(row=r_main, column=27).value)
        if op != mp: ok = False; print(f"  ✗ 画{gi+1} r{r_out} parSKU: OUT={op} vs MAIN={mp}")
print(f"  {'✓' if ok else '✗'} 普文件原 11 行 parSKU 全部保留")

print("\n" + "=" * 70)
print("全部检查完成")
print("=" * 70)

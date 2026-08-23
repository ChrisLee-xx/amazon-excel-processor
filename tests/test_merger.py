"""合并模块测试 (3 文件 API: 主+木+金)"""

import random
import pytest
from openpyxl import Workbook, load_workbook

from amazon_excel_processor.excel_io import DATA_START_ROW, HEADER_ROW, group_rows
from amazon_excel_processor.name_normalizer import VARIANT_LABELS_21
from amazon_excel_processor.merger import (
    identify_file_role,
    identify_main_file,  # 向后兼容
    index_groups_by_name,
    merge_one_painting,
    rewrite_sku,
    write_parent_sku_formulas,
    build_sku_prefix,
    MAIN_GROUP_SIZE,
    VARIANT_GROUP_SIZE,
)


# ===== 测试夹具 =====

def _create_main_workbook(paintings):
    """模拟"普文件"(新格式): N 画各 11 行 (1 parent + 5 Frame + 5 Unframe). 返回 (wb, col_map)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = {
        1: "SKU", 4: "Parentage Level", 5: "Parent SKU", 6: "Variation Theme Name",
        7: "Item Name", 46: "Style", 55: "Color", 56: "Size",
        124: "Item Length Longer Edge", 126: "Item Width Shorter Edge",
        147: "Item Weight", 154: "List Price",
        182: "Your Price USD (Sell on Amazon, US)",
        191: "Your Price USD (Amazon Business (B2B), US)",
    }
    for c, h in headers.items():
        ws.cell(row=HEADER_ROW, column=c).value = h
    row = DATA_START_ROW
    for title in paintings:
        ws.cell(row=row, column=7).value = title
        ws.cell(row=row, column=1).value = f"SKU-{row}"
        ws.cell(row=row, column=4).value = "Parent"
        row += 1
        for i in range(10):
            ws.cell(row=row, column=7).value = (
                f"{title} Frame-style 12x18inch" if i < 5
                else f"{title} Unframe-style 12x18inch"
            )
            ws.cell(row=row, column=1).value = f"SKU-{row}"
            ws.cell(row=row, column=4).value = "Child"
            row += 1
    col_map = {h: c for c, h in headers.items()}
    return wb, col_map


def _create_variant_workbook(paintings, role="wood", shuffled=False):
    """模拟"木框/金框文件"(新格式): N 画各 1 个 6 行 group.

    parent 行 Item Name 不带 style 关键字 (与真实文件一致).
    children 用 {role} 后缀区分 (便于测试时验证来源).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    for c, h in {1: "SKU", 4: "Parentage Level", 7: "Item Name"}.items():
        ws.cell(row=HEADER_ROW, column=c).value = h
    items = list(paintings)
    if shuffled:
        random.seed(42)
        random.shuffle(items)
    row = DATA_START_ROW
    for title in items:
        ws.cell(row=row, column=7).value = title
        ws.cell(row=row, column=1).value = f"{role.upper()}-{row}"
        ws.cell(row=row, column=4).value = "Parent"
        row += 1
        for i in range(5):
            ws.cell(row=row, column=7).value = f"{title} {role} size-{i}"
            ws.cell(row=row, column=1).value = f"{role.upper()}-{row}"
            ws.cell(row=row, column=4).value = "Child"
            row += 1
    return wb


# ===== identify_file_role =====

class TestIdentifyFileRole:
    def test_11_rows_is_main(self):
        wb, _ = _create_main_workbook(["A", "B"])
        groups = group_rows(wb.active, group_size=MAIN_GROUP_SIZE)
        role, _ = identify_file_role(groups)
        assert role == "main"

    def test_6_rows_is_variant(self):
        wb = _create_variant_workbook(["A", "B"], role="wood")
        groups = group_rows(wb.active, group_size=VARIANT_GROUP_SIZE)
        role, _ = identify_file_role(groups)
        assert role == "variant"

    def test_backward_compat_identify_main_file(self):
        """旧 API identify_main_file 仍可用"""
        wb, _ = _create_main_workbook(["A"])
        groups = group_rows(wb.active, group_size=MAIN_GROUP_SIZE)
        is_main, is_gold = identify_main_file(groups)
        assert is_main is True
        assert is_gold is False


# ===== index_groups_by_name =====

class TestIndexGroupsByName:
    def test_index_4_paintings(self):
        wb = _create_variant_workbook(["Art A", "Art B", "Art C", "Art D"], role="wood")
        ws = wb.active
        groups = group_rows(ws, group_size=VARIANT_GROUP_SIZE)
        by_name = index_groups_by_name(ws, groups)
        assert len(by_name) == 4

    def test_duplicate_allowed_by_order(self):
        """同 base name 出现多次允许, 按顺序配对 (普第 N 次 ↔ 木第 N 次)"""
        wb = _create_variant_workbook(["Art A", "Art A"], role="wood")
        ws = wb.active
        groups = group_rows(ws, group_size=VARIANT_GROUP_SIZE)
        by_name = index_groups_by_name(ws, groups)
        assert len(by_name) == 1
        assert len(by_name["art a"]) == 2  # 2 个 group, 按出现顺序


# ===== merge_one_painting =====

def _setup_merge_one_painting(painting="Art A"):
    """构造单画的 main + wood + gold 测试数据, 返回所需参数."""
    main_wb, main_col_map = _create_main_workbook([painting])
    main_ws = main_wb.active
    main_groups = group_rows(main_ws, group_size=MAIN_GROUP_SIZE)

    wood_ws = _create_variant_workbook([painting], role="wood").active
    wood_groups = group_rows(wood_ws, group_size=VARIANT_GROUP_SIZE)

    gold_ws = _create_variant_workbook([painting], role="gold").active
    gold_groups = group_rows(gold_ws, group_size=VARIANT_GROUP_SIZE)

    max_col = max(main_ws.max_column, wood_ws.max_column, gold_ws.max_column)
    main_snapshots = [
        {c: main_ws.cell(row=r, column=c).value for c in range(1, max_col + 1)}
        for r in main_groups[0]
    ]
    return {
        "main_ws": main_ws,
        "main_col_map": main_col_map,
        "main_snapshots": main_snapshots,
        "wood_group": wood_groups[0],
        "gold_group": gold_groups[0],
        "wood_ws": wood_ws,
        "gold_ws": gold_ws,
        "max_col": max_col,
    }


class TestMergeOnePainting:
    def test_output_21_rows(self):
        s = _setup_merge_one_painting()
        merged = merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
        )
        assert len(merged) == 21
        assert merged[0] == 4
        assert merged[-1] == 24

    def test_output_color_sequence(self):
        s = _setup_merge_one_painting()
        merged = merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
        )
        colors = [s["main_ws"].cell(row=r, column=55).value for r in merged]
        # merged[0]=r4 parent, [1-5]=Frame, [6-10]=Unframe, [11-15]=Wood, [16-20]=Gold
        assert colors[0] in (None, "")
        assert colors[1] == "Frame-style"
        assert colors[5] == "Frame-style"
        assert colors[6] == "Unframe-style"
        assert colors[10] == "Unframe-style"
        assert colors[11] == "Vintage Wood Grain Frame-style"
        assert colors[15] == "Vintage Wood Grain Frame-style"
        assert colors[16] == "Vintage Ornate Gold Frame-style"
        assert colors[20] == "Vintage Ornate Gold Frame-style"

    def test_wood_gold_match_frame_size(self):
        s = _setup_merge_one_painting()
        merged = merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
        )
        # Frame r5-r9, Wood r15-r19, Gold r20-r24
        # 注意: Color/Weight 不同 style 本来就不同, 不比较; 只比较 Size/Length/Width (物理尺寸一致)
        for rf, rw, rg in [(5, 15, 20), (6, 16, 21), (7, 17, 22), (8, 18, 23), (9, 19, 24)]:
            ws = s["main_ws"]
            assert ws.cell(row=rf, column=56).value == ws.cell(row=rw, column=56).value
            assert ws.cell(row=rf, column=56).value == ws.cell(row=rg, column=56).value
            assert ws.cell(row=rf, column=124).value == ws.cell(row=rw, column=124).value
            assert ws.cell(row=rf, column=124).value == ws.cell(row=rg, column=124).value
            assert ws.cell(row=rf, column=126).value == ws.cell(row=rw, column=126).value
            assert ws.cell(row=rf, column=126).value == ws.cell(row=rg, column=126).value

    def test_parentage_relationship_type(self):
        s = _setup_merge_one_painting()
        merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
        )
        ws = s["main_ws"]
        assert ws.cell(row=4, column=4).value == "Parent"
        assert ws.cell(row=5, column=4).value == "Child"
        assert ws.cell(row=24, column=4).value == "Child"
        # 新格式: Variation Theme Name (col6) = "COLOR/SIZE"
        assert ws.cell(row=5, column=6).value == "COLOR/SIZE"
        assert ws.cell(row=24, column=6).value == "COLOR/SIZE"

    def test_wood_image_from_wood_file(self):
        """Wood 行的 Image URL 必须来自 wood 文件, Gold 行的来自 gold 文件.

        这是 3 文件方案的核心: 用户在 GUI 中明确指定 wood/gold, 不再依赖出现顺序.
        """
        s = _setup_merge_one_painting()
        merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
        )
        ws = s["main_ws"]
        # Wood r15 (第 1 个 wood child) 的 SKU 应来自 wood 文件
        # 测试夹具里 wood SKU 是 "WOOD-{row}", gold SKU 是 "GOLD-{row}"
        # 但 merge_one_painting 后 SKU 没被重写 (rewrite_sku 是单独步骤)
        # 所以 r15 col2 应该是 wood 文件第 1 个 child 的 SKU
        wood_child_sku = ws.cell(row=15, column=1).value
        gold_child_sku = ws.cell(row=20, column=1).value
        # wood 文件第 1 个 group 的第 1 个 child 是 r5 (parent r4, child r5-r9)
        assert wood_child_sku == "WOOD-9"
        assert gold_child_sku == "GOLD-9"

    def test_old_variant_mode_preserves_main_rows(self):
        """老品补充变体模式: 普文件原 11 行 (r4-r14) 完全不动, 仅 Wood/Gold 行处理"""
        s = _setup_merge_one_painting()
        main_ws = s["main_ws"]
        main_col_map = s["main_col_map"]
        main_snapshots = s["main_snapshots"]

        # 先记录普文件原 11 行的所有列值 (从快照)
        original_values = {}
        for i in range(11):
            for c, v in main_snapshots[i].items():
                original_values[(i, c)] = v

        # 用老品补充模式合并
        merge_one_painting(
            main_snapshots=main_snapshots,
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=main_ws,
            col_map=main_col_map,
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
            mode="old_variant",
        )

        # 验证 r4-r14 (普文件原 11 行) 的所有列值没变
        for i in range(11):
            r = 4 + i
            for c in range(1, s["max_col"] + 1):
                orig = original_values[(i, c)]
                now = main_ws.cell(row=r, column=c).value
                # 空字符串和 None 视为相同
                orig_e = orig if orig not in (None, "") else ""
                now_e = now if now not in (None, "") else ""
                assert orig_e == now_e, f"r{r} col{c}: 原值={orig} 现值={now} (老品补充不应修改普文件原行)"

    def test_old_variant_mode_wood_gold_color_price(self):
        """老品补充变体模式: Wood/Gold 行的 Color 和 Price 仍正确填充"""
        s = _setup_merge_one_painting()
        main_ws = s["main_ws"]
        merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=main_ws,
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
            mode="old_variant",
        )
        # Wood r15-r19
        for i, r in enumerate(range(15, 20)):
            assert main_ws.cell(row=r, column=55).value == "Vintage Wood Grain Frame-style"
            assert main_ws.cell(row=r, column=154).value == [26.9, 39.9, 59.9, 99.9, 129.9][i]
        # Gold r20-r24
        for i, r in enumerate(range(20, 25)):
            assert main_ws.cell(row=r, column=55).value == "Vintage Ornate Gold Frame-style"
            assert main_ws.cell(row=r, column=154).value == [26.9, 39.9, 59.9, 99.9, 129.9][i]

    def test_old_variant_mode_wood_gold_three_price_columns(self):
        """老品补充变体模式: Wood/Gold 行的 3 个价格列都按尺寸序列填充。

        价格列: List Price (col154) / Your Price (col182) / B2B (col191)。
        木金价格按 size 从小到大: 26.9 / 39.9 / 59.9 / 99.9 / 129.9。
        """
        s = _setup_merge_one_painting()
        main_ws = s["main_ws"]
        merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=main_ws,
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
            mode="old_variant",
        )
        expected = [26.9, 39.9, 59.9, 99.9, 129.9]
        # Wood r15-r19 + Gold r20-r24, 3 个价格列全部等于价格序列
        for i, r in enumerate(list(range(15, 20)) + list(range(20, 25))):
            for c in [154, 182, 191]:
                assert main_ws.cell(row=r, column=c).value == expected[i % 5], \
                    f"r{r} col{c}: 期望 {expected[i % 5]}, 实际 {main_ws.cell(row=r, column=c).value}"

    def test_new_mode_wood_gold_three_price_columns(self):
        """新品上架模式: Wood/Gold 行的 3 个价格列都按尺寸序列填充。"""
        s = _setup_merge_one_painting()
        main_ws = s["main_ws"]
        merged = merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=4,
            output_ws=main_ws,
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
            mode="new",
        )
        wood_gold_rows = merged[11:]  # Wood×5 + Gold×5
        expected = [26.9, 39.9, 59.9, 99.9, 129.9]
        for i, r in enumerate(wood_gold_rows):
            for c in [154, 182, 191]:
                assert main_ws.cell(row=r, column=c).value == expected[i % 5], \
                    f"r{r} col{c}: 期望 {expected[i % 5]}, 实际 {main_ws.cell(row=r, column=c).value}"
        # 普通 Frame/Unframe 行的 3 个价格列也同步填充 (frame 前 5, unframe 后 5)
        frame_prices = [19.9, 29.9, 45, 75, 99]
        unframe_prices = [11.9, 14.9, 19.9, 24.9, 34.9]
        for i, r in enumerate(merged[1:11]):
            expected_p = frame_prices[i] if i < 5 else unframe_prices[i - 5]
            for c in [154, 182, 191]:
                assert main_ws.cell(row=r, column=c).value == expected_p

    def test_parent_row_style_copied_from_e8(self):
        """parent 行的 Parent SKU (E 列) 样式应复制自模板 E8 (深色填充)."""
        from openpyxl.styles import PatternFill
        from amazon_excel_processor.excel_io import DATA_START_ROW
        s = _setup_merge_one_painting()
        main_ws = s["main_ws"]
        # 给模板 r=DATA_START_ROW (模拟 E8) 设置一个明显的填充
        template_e = main_ws.cell(row=DATA_START_ROW, column=5)
        template_e.fill = PatternFill(
            patternType='solid', fgColor='FF632523'
        )
        merge_one_painting(
            main_snapshots=s["main_snapshots"],
            wood_group=s["wood_group"],
            gold_group=s["gold_group"],
            output_start_row=DATA_START_ROW,
            output_ws=main_ws,
            col_map=s["main_col_map"],
            wood_ws=s["wood_ws"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
        )
        # parent 行 (output_start_row) 的 E 列 fill 应与原模板一致 (FG=FF632523)
        parent_e = main_ws.cell(row=DATA_START_ROW, column=5)
        assert parent_e.fill.fgColor.rgb == 'FF632523'


# ===== rewrite_sku =====

class TestRewriteSku:
    def test_continuous_numbering_new_mode(self):
        """新品上架: 父体={prefix}-N, 普通子体={prefix}P-N, 木框={prefix}W-N, 金框={prefix}J-N, 四套独立编号"""
        wb = Workbook()
        ws = wb.active
        groups = [list(range(4, 25)), list(range(25, 46))]
        rewrite_sku(ws, groups, prefix="HM725", mode="new", has_wood=True, has_gold=True)
        # group 1: parent = HM725-1
        assert ws.cell(row=4, column=1).value == "HM725-1"
        # group 1: 普通子体 (r5-r14) = HM725P-1 到 HM725P-10
        assert ws.cell(row=5, column=1).value == "HM725P-1"
        assert ws.cell(row=14, column=1).value == "HM725P-10"
        # group 1: 木框子体 (r15-r19) = HM725M-1 到 HM725M-5
        assert ws.cell(row=15, column=1).value == "HM725M-1"
        assert ws.cell(row=19, column=1).value == "HM725M-5"
        # group 1: 金框子体 (r20-r24) = HM725J-1 到 HM725J-5
        assert ws.cell(row=20, column=1).value == "HM725J-1"
        assert ws.cell(row=24, column=1).value == "HM725J-5"
        # group 2: parent = HM725-2
        assert ws.cell(row=25, column=1).value == "HM725-2"
        # group 2: 普通子体 = HM725P-11 到 HM725P-20
        assert ws.cell(row=26, column=1).value == "HM725P-11"
        assert ws.cell(row=35, column=1).value == "HM725P-20"
        # group 2: 木框子体 = HM725M-6 到 HM725M-10
        assert ws.cell(row=36, column=1).value == "HM725M-6"
        assert ws.cell(row=40, column=1).value == "HM725M-10"
        # group 2: 金框子体 = HM725J-6 到 HM725J-10
        assert ws.cell(row=41, column=1).value == "HM725J-6"
        assert ws.cell(row=45, column=1).value == "HM725J-10"

    def test_single_group_new_mode(self):
        wb = Workbook()
        ws = wb.active
        groups = [list(range(4, 25))]
        rewrite_sku(ws, groups, prefix="AB", mode="new", has_wood=True, has_gold=True)
        # parent
        assert ws.cell(row=4, column=1).value == "AB-1"
        # 普通子体
        assert ws.cell(row=5, column=1).value == "ABP-1"
        assert ws.cell(row=14, column=1).value == "ABP-10"
        # 木框子体 (r15-r19)
        assert ws.cell(row=15, column=1).value == "ABM-1"
        assert ws.cell(row=19, column=1).value == "ABM-5"
        # 金框子体 (r20-r24)
        assert ws.cell(row=20, column=1).value == "ABJ-1"
        assert ws.cell(row=24, column=1).value == "ABJ-5"

    def test_wood_only_uses_M_suffix(self):
        """只有木框 (无金框) → 木框用 M 后缀, 无 J 行"""
        wb = Workbook()
        ws = wb.active
        groups = [list(range(4, 20))]  # 16 行: parent + 10 普通 + 5 木
        rewrite_sku(ws, groups, prefix="T", mode="new", has_wood=True, has_gold=False)
        assert ws.cell(row=4, column=1).value == "T-1"
        assert ws.cell(row=5, column=1).value == "TP-1"
        assert ws.cell(row=14, column=1).value == "TP-10"
        assert ws.cell(row=15, column=1).value == "TM-1"
        assert ws.cell(row=19, column=1).value == "TM-5"

    def test_gold_only_uses_J_suffix(self):
        """只有金框 (无木框) → 金框用 J 后缀, 无 W 行"""
        wb = Workbook()
        ws = wb.active
        groups = [list(range(4, 20))]  # 16 行: parent + 10 普通 + 5 金
        rewrite_sku(ws, groups, prefix="T", mode="new", has_wood=False, has_gold=True)
        assert ws.cell(row=4, column=1).value == "T-1"
        assert ws.cell(row=5, column=1).value == "TP-1"
        assert ws.cell(row=14, column=1).value == "TP-10"
        assert ws.cell(row=15, column=1).value == "TJ-1"
        assert ws.cell(row=19, column=1).value == "TJ-5"

    def test_old_variant_mode_preserves_main_sku(self):
        """老品补充变体: group[0:11] (普文件原 11 行) SKU 保留, 木框=W, 金框=J"""
        wb = Workbook()
        ws = wb.active
        groups = [list(range(4, 25))]
        # 预设普文件原 11 行的 SKU (group[0:11] = r4-r14)
        for i, r in enumerate(groups[0][:11]):
            ws.cell(row=r, column=1).value = f"OLD-{i+1}"
        rewrite_sku(ws, groups, prefix="NEW", mode="old_variant", has_wood=True, has_gold=True)
        # r4-r14 保留原 SKU
        assert ws.cell(row=4, column=1).value == "OLD-1"
        assert ws.cell(row=14, column=1).value == "OLD-11"
        # r15-r19 (Wood) = NEWM-1 到 NEWM-5
        assert ws.cell(row=15, column=1).value == "NEWM-1"
        assert ws.cell(row=19, column=1).value == "NEWM-5"
        # r20-r24 (Gold) = NEWJ-1 到 NEWJ-5
        assert ws.cell(row=20, column=1).value == "NEWJ-1"
        assert ws.cell(row=24, column=1).value == "NEWJ-5"

    def test_old_variant_mode_multi_groups_continuous(self):
        """老品补充变体: 多 group 时 Wood/Gold 各自跨 group 连续编号"""
        wb = Workbook()
        ws = wb.active
        groups = [list(range(4, 25)), list(range(25, 46))]
        rewrite_sku(ws, groups, prefix="NEW", mode="old_variant", has_wood=True, has_gold=True)
        # group 1: r15-r19 (Wood) = NEWM-1 到 NEWM-5
        assert ws.cell(row=15, column=1).value == "NEWM-1"
        assert ws.cell(row=19, column=1).value == "NEWM-5"
        # group 1: r20-r24 (Gold) = NEWJ-1 到 NEWJ-5
        assert ws.cell(row=20, column=1).value == "NEWJ-1"
        assert ws.cell(row=24, column=1).value == "NEWJ-5"
        # group 2: r36-r40 (Wood) = NEWM-6 到 NEWM-10
        assert ws.cell(row=36, column=1).value == "NEWM-6"
        assert ws.cell(row=40, column=1).value == "NEWM-10"
        # group 2: r41-r45 (Gold) = NEWJ-6 到 NEWJ-10
        assert ws.cell(row=41, column=1).value == "NEWJ-6"
        assert ws.cell(row=45, column=1).value == "NEWJ-10"


# ===== write_parent_sku_formulas =====

class TestWriteParentSkuFormulas:
    def test_first_child_uses_b_parent_new_mode(self):
        """新品上架: parent 清空, 第 1 child =B4, 后续 =AA{prev}"""
        wb = Workbook()
        ws = wb.active
        rows = list(range(4, 25))
        write_parent_sku_formulas(ws, [rows], mode="new")
        assert ws.cell(row=4, column=5).value in (None, "")
        assert ws.cell(row=5, column=5).value == "=A4"
        assert ws.cell(row=6, column=5).value == "=E5"
        assert ws.cell(row=7, column=5).value == "=E6"
        assert ws.cell(row=24, column=5).value == "=E23"

    def test_multiple_groups_new_mode(self):
        wb = Workbook()
        ws = wb.active
        write_parent_sku_formulas(ws, [list(range(4, 25)), list(range(25, 46))], mode="new")
        assert ws.cell(row=5, column=5).value == "=A4"
        assert ws.cell(row=6, column=5).value == "=E5"
        assert ws.cell(row=26, column=5).value == "=A25"
        assert ws.cell(row=27, column=5).value == "=E26"

    def test_old_variant_mode_preserves_main_formulas(self):
        """老品补充变体: group[0:11] parent SKU 公式保留, group[11:21] 从 =AA{group[10]} 开始"""
        wb = Workbook()
        ws = wb.active
        rows = list(range(4, 25))
        # 预设普文件原 11 行的 parent SKU 公式
        ws.cell(row=4, column=5).value = None       # parent
        ws.cell(row=5, column=5).value = "=A4"      # Frame 1
        ws.cell(row=6, column=5).value = "=E5"
        ws.cell(row=14, column=5).value = "=E13"   # Unframe 5 (最后一个)
        write_parent_sku_formulas(ws, [rows], mode="old_variant")
        # r4-r14 保留
        assert ws.cell(row=4, column=5).value is None
        assert ws.cell(row=5, column=5).value == "=A4"
        assert ws.cell(row=14, column=5).value == "=E13"
        # r15 (Wood 1) = =AA14 (引用 r14 Unframe 最后一个)
        assert ws.cell(row=15, column=5).value == "=E14"
        # r16 = =AA15
        assert ws.cell(row=16, column=5).value == "=E15"
        # r24 (Gold 5) = =AA23
        assert ws.cell(row=24, column=5).value == "=E23"


# ===== build_sku_prefix =====

class TestBuildSkuPrefix:
    def test_single_string(self):
        assert build_sku_prefix("HM725") == "HM725"

    def test_strip_whitespace(self):
        assert build_sku_prefix("  HM725  ") == "HM725"

    def test_complex_prefix(self):
        assert build_sku_prefix("AB2026风景") == "AB2026风景"


# ===== 木/金可选 (动态行数 11/16/21) =====

class TestOptionalVariants:
    """木框/金框独立可选: 输出 11/16/21 行。"""

    def test_wood_only_16_rows(self):
        """只提供木框 (无金框) → 16 行, 木 color 在 index 11-15, 无金行。"""
        s = _setup_merge_one_painting()
        merged = merge_one_painting(
            main_snapshots=s["main_snapshots"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_group=s["wood_group"],
            wood_ws=s["wood_ws"],
            gold_group=None,
            gold_ws=None,
            max_col=s["max_col"],
        )
        assert len(merged) == 16
        assert merged[0] == 4
        assert merged[-1] == 19
        ws = s["main_ws"]
        colors = [ws.cell(row=r, column=55).value for r in merged]
        assert colors[0] in (None, "")
        assert colors[1] == "Frame-style"
        assert colors[10] == "Unframe-style"
        for i in range(11, 16):
            assert colors[i] == "Vintage Wood Grain Frame-style"
        # 3 个价格列 (List Price / Your Price / B2B) 都按尺寸序列填充
        for c in [154, 182, 191]:
            prices = [ws.cell(row=r, column=c).value for r in merged]
            assert prices[11:16] == [26.9, 39.9, 59.9, 99.9, 129.9]

    def test_gold_only_16_rows(self):
        """只提供金框 (无木框) → 16 行, 金 color 紧随 main 占 index 11-15。"""
        s = _setup_merge_one_painting()
        merged = merge_one_painting(
            main_snapshots=s["main_snapshots"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_group=None,
            wood_ws=None,
            gold_group=s["gold_group"],
            gold_ws=s["gold_ws"],
            max_col=s["max_col"],
        )
        assert len(merged) == 16
        ws = s["main_ws"]
        colors = [ws.cell(row=r, column=55).value for r in merged]
        for i in range(11, 16):
            assert colors[i] == "Vintage Ornate Gold Frame-style"
        # 金行数据来自 gold 文件 (gold 第 1 个 child SKU = GOLD-9, 新格式 DATA_START_ROW=8)
        assert ws.cell(row=merged[11], column=1).value == "GOLD-9"

    def test_main_only_new_mode_11_rows(self):
        """无木无金, new 模式 → 11 行 (parent + Frame×5 + Unframe×5)。"""
        s = _setup_merge_one_painting()
        merged = merge_one_painting(
            main_snapshots=s["main_snapshots"],
            output_start_row=4,
            output_ws=s["main_ws"],
            col_map=s["main_col_map"],
            wood_group=None,
            wood_ws=None,
            gold_group=None,
            gold_ws=None,
            max_col=s["max_col"],
            mode="new",
        )
        assert len(merged) == 11
        assert merged[-1] == 14

    def test_old_variant_requires_variant(self):
        """old_variant 模式无变体 → 报错。"""
        s = _setup_merge_one_painting()
        with pytest.raises(ValueError, match="至少一个木框或金框"):
            merge_one_painting(
                main_snapshots=s["main_snapshots"],
                output_start_row=4,
                output_ws=s["main_ws"],
                col_map=s["main_col_map"],
                wood_group=None,
                gold_group=None,
                mode="old_variant",
            )

    def test_old_variant_wood_only_preserves_main(self):
        """old_variant + 只木 → main 11 行不动, 木 5 行处理, 无金行。"""
        s = _setup_merge_one_painting()
        main_ws = s["main_ws"]
        main_snapshots = s["main_snapshots"]
        original = {}
        for i in range(11):
            for c, v in main_snapshots[i].items():
                original[(i, c)] = v
        merge_one_painting(
            main_snapshots=main_snapshots,
            output_start_row=4,
            output_ws=main_ws,
            col_map=s["main_col_map"],
            wood_group=s["wood_group"],
            wood_ws=s["wood_ws"],
            gold_group=None,
            gold_ws=None,
            max_col=s["max_col"],
            mode="old_variant",
        )
        # main 11 行 (r4-r14) 完全不变
        for i in range(11):
            r = 4 + i
            for c in range(1, s["max_col"] + 1):
                orig = original[(i, c)]
                now = main_ws.cell(row=r, column=c).value
                orig_e = orig if orig not in (None, "") else ""
                now_e = now if now not in (None, "") else ""
                assert orig_e == now_e, f"r{r} col{c}: 原值={orig} 现值={now}"
        # 木行 r15-r19
        for i, r in enumerate(range(15, 20)):
            assert main_ws.cell(row=r, column=55).value == "Vintage Wood Grain Frame-style"
            assert main_ws.cell(row=r, column=154).value == [26.9, 39.9, 59.9, 99.9, 129.9][i]
        # 无金行
        assert main_ws.cell(row=20, column=55).value in (None, "")


# ===== merge_files 集成 (木/金可选) =====

class TestMergeFilesOptional:
    """merge_files 端到端: 木/金可选 + 配对错误处理。"""

    def test_gold_only_missing_painting_raises_cleanly(self, tmp_path):
        """金 only 且金文件缺某画 → _raise_pairing_error 干净触发 (无 None 崩溃)。"""
        from amazon_excel_processor.merger import merge_files
        main_wb, _ = _create_main_workbook(["Art A", "Art B"])
        gold_wb = _create_variant_workbook(["Art A"], role="gold")  # 缺 Art B
        main_p = tmp_path / "main.xlsx"
        gold_p = tmp_path / "gold.xlsx"
        main_wb.save(str(main_p))
        gold_wb.save(str(gold_p))
        with pytest.raises(ValueError) as excinfo:
            merge_files(main_path=main_p, wood_path=None, gold_path=gold_p,
                        sku_prefix="T", mode="new")
        msg = str(excinfo.value)
        assert "配对失败" in msg
        assert "未提供" in msg  # 木框未提供
        assert "金框" in msg

    def test_wood_only_merge_files_16_rows(self, tmp_path):
        """merge_files 木 only → 输出 16 行/组, 无金行。"""
        from amazon_excel_processor.merger import merge_files
        from openpyxl import load_workbook as _lw
        main_wb, _ = _create_main_workbook(["Art A"])
        wood_wb = _create_variant_workbook(["Art A"], role="wood")
        main_p = tmp_path / "main.xlsx"
        wood_p = tmp_path / "wood.xlsx"
        main_wb.save(str(main_p))
        wood_wb.save(str(wood_p))
        out = merge_files(main_path=main_p, wood_path=wood_p, gold_path=None,
                          sku_prefix="T", mode="new")
        ws = _lw(str(out))["Template"]
        # parent + 10 main + 5 wood = 16 行 (新格式 DATA_START_ROW=8, r8-r23)
        assert ws.cell(row=8, column=4).value == "Parent"
        assert ws.cell(row=9, column=4).value == "Child"
        assert ws.cell(row=23, column=55).value == "Vintage Wood Grain Frame-style"
        assert ws.cell(row=24, column=55).value in (None, "")  # 无金行

    def test_pairing_error_shows_product_name(self, tmp_path):
        """配对失败时错误信息应包含原始 Product Name (不是空)。

        回归: merge_files 清空 main_ws 数据区后, 配对失败的 Product Name
        必须从快照读取, 否则错误信息中 Product Name 为空。
        """
        from amazon_excel_processor.merger import merge_files
        main_wb, _ = _create_main_workbook(["Art A", "Sunset Beach"])
        gold_wb = _create_variant_workbook(["Art A"], role="gold")  # 缺 Sunset Beach
        main_p = tmp_path / "main.xlsx"
        gold_p = tmp_path / "gold.xlsx"
        main_wb.save(str(main_p))
        gold_wb.save(str(gold_p))
        with pytest.raises(ValueError) as excinfo:
            merge_files(main_path=main_p, wood_path=None, gold_path=gold_p,
                        sku_prefix="T", mode="new")
        msg = str(excinfo.value)
        assert "Sunset Beach" in msg  # 原始 Product Name 应出现, 不是空


# ===== Search Terms 清理 (new 模式) =====

class TestSearchTermsCleaning:
    """new 模式 (fill_group_merged) 应清理 Search Terms 下划线。"""

    def test_new_mode_cleans_search_terms(self):
        """new 模式下 Search Terms 的下划线应替换为空格。"""
        from amazon_excel_processor.field_filler import fill_group_merged, build_active_styles
        wb = Workbook()
        ws = wb.active
        col_map = {"Search Terms": 200, "Color": 55}
        rows = list(range(8, 19))  # 11 行
        for r in rows:
            ws.cell(row=r, column=200).value = "art_print_wall_decor"
        fill_group_merged(ws, rows, col_map, "3:2", build_active_styles(False, False))
        for r in rows:
            assert ws.cell(row=r, column=200).value == "art print wall decor"


# ===== 向后兼容: 动态序列 == 21 行常量 =====

class TestDynamicSequencesBackwardCompat:
    """build_active_styles(True,True) 动态序列必须与现有 *_21 常量逐元素相等。"""

    def test_sequences_match_21_constants(self):
        from amazon_excel_processor.field_filler import (
            build_active_styles, _build_sequences,
            COLOR_SEQUENCE_21, SIZE_MAP_SEQUENCE_21, SIZE_32_21,
            LENGTH_32_21, WIDTH_32_21, WEIGHT_SEQUENCE_21, PRICE_SEQUENCE_21,
            PACKAGE_LENGTH_21, PACKAGE_WIDTH_21, PACKAGE_HEIGHT_21, PACKAGE_WEIGHT_21,
        )
        seqs = _build_sequences(build_active_styles(True, True))
        assert seqs["color"] == COLOR_SEQUENCE_21
        assert seqs["size_map"] == SIZE_MAP_SEQUENCE_21
        assert seqs["size_32"] == SIZE_32_21
        assert seqs["length"] == LENGTH_32_21
        assert seqs["width"] == WIDTH_32_21
        assert seqs["weight"] == WEIGHT_SEQUENCE_21
        assert seqs["price"] == PRICE_SEQUENCE_21
        assert seqs["labels"][1:] == VARIANT_LABELS_21[1:]
        assert seqs["package_length"] == PACKAGE_LENGTH_21
        assert seqs["package_width"] == PACKAGE_WIDTH_21
        assert seqs["package_height"] == PACKAGE_HEIGHT_21
        assert seqs["package_weight"] == PACKAGE_WEIGHT_21


# ===== Style 列保留原始值 (不被 Color 覆盖) =====

class TestStyleColumnPreserved:
    """fill_group_merged / _fill_variant_fields 不应覆盖 Style 列, 只填 Color 列。"""

    def test_fill_group_merged_keeps_style(self):
        """new 模式: Style 列保留原始值, Color 列填 style 标签。"""
        from amazon_excel_processor.field_filler import fill_group_merged, build_active_styles
        wb = Workbook()
        ws = wb.active
        col_map = {"Color": 55, "Style": 46}
        rows = list(range(8, 19))  # 11 行
        # 预设 Style 列原始值 (模拟用户手填)
        original_styles = ["orig-parent", "orig-f1", "orig-f2", "orig-f3", "orig-f4", "orig-f5",
                           "orig-u1", "orig-u2", "orig-u3", "orig-u4", "orig-u5"]
        for i, r in enumerate(rows):
            ws.cell(row=r, column=46).value = original_styles[i]
        fill_group_merged(ws, rows, col_map, "3:2", build_active_styles(False, False))
        # Style 列应保留原始值
        for i, r in enumerate(rows):
            assert ws.cell(row=r, column=46).value == original_styles[i], \
                f"row {r} Style 列被覆盖"
        # Color 列应被填充 (parent 空, 子体有标签)
        assert ws.cell(row=rows[0], column=55).value == ""
        assert ws.cell(row=rows[1], column=55).value == "Frame-style"
        assert ws.cell(row=rows[6], column=55).value == "Unframe-style"


# ===== 单文件模式 SKU 重写 (无木金 J 后缀) =====

class TestSingleFileSkuRewrite:
    """单文件 (11 行/组) SKU 命名: 父体 prefix-N, 普通子体 prefixP-N。"""

    def test_single_file_sku_naming(self):
        """单文件 mode=new: parent=prefix-1, 子体=prefixP-1..prefixP-10。"""
        from amazon_excel_processor.merger import rewrite_sku, write_parent_sku_formulas
        wb = Workbook()
        ws = wb.active
        ws.cell(row=HEADER_ROW, column=1).value = "SKU"
        ws.cell(row=HEADER_ROW, column=5).value = "Parent SKU"
        # 2 组 × 11 行 = 22 行
        groups = [list(range(8, 19)), list(range(19, 30))]
        for g in groups:
            for r in g:
                ws.cell(row=r, column=1).value = f"OLD-{r}"
        rewrite_sku(ws, groups, "XL810Z", sku_col=1, mode="new")
        # 第 1 组
        assert ws.cell(row=8, column=1).value == "XL810Z-1"      # parent
        assert ws.cell(row=9, column=1).value == "XL810ZP-1"     # 子体 1
        assert ws.cell(row=18, column=1).value == "XL810ZP-10"   # 子体 10
        # 第 2 组
        assert ws.cell(row=19, column=1).value == "XL810Z-2"     # parent
        assert ws.cell(row=20, column=1).value == "XL810ZP-11"   # 子体 11
        assert ws.cell(row=29, column=1).value == "XL810ZP-20"   # 子体 20
        # 无 J 后缀 (单文件没有木金)
        for g in groups:
            for r in g[1:]:
                assert "J" not in str(ws.cell(row=r, column=1).value)

    def test_single_file_parent_sku_formulas(self):
        """单文件 mode=new: parent 行 Parent SKU 清空, 子体用公式引用。"""
        from amazon_excel_processor.merger import write_parent_sku_formulas
        wb = Workbook()
        ws = wb.active
        ws.cell(row=HEADER_ROW, column=1).value = "SKU"
        ws.cell(row=HEADER_ROW, column=5).value = "Parent SKU"
        groups = [list(range(8, 19))]
        # parent
        ws.cell(row=8, column=1).value = "XL810Z-1"
        # 子体
        for i, r in enumerate(groups[0][1:], 1):
            ws.cell(row=r, column=1).value = f"XL810ZP-{i}"
        write_parent_sku_formulas(ws, groups, parent_sku_col=5, seller_sku_col=1, mode="new")
        # parent 行 Parent SKU 为空
        assert ws.cell(row=8, column=5).value is None
        # 第 1 个子体 = =A8 (引用 parent 的 Seller SKU)
        assert ws.cell(row=9, column=5).value == "=A8"
        # 后续子体 = =E{prev} (引用上一行 Parent SKU, 链式)
        assert ws.cell(row=10, column=5).value == "=E9"
        assert ws.cell(row=18, column=5).value == "=E17"


# ===== old_parent 模式无 wood/gold 报错 =====

class TestOldParentNoVariantRaises:
    """old_parent 模式 (与 old_variant 一致) 无 wood/gold 时应报错。"""

    def test_old_parent_no_variant_raises(self, tmp_path):
        from amazon_excel_processor.merger import merge_files
        main_wb, _ = _create_main_workbook(["Art A"])
        main_p = tmp_path / "main.xlsx"
        main_wb.save(str(main_p))
        with pytest.raises(ValueError, match="老品模式"):
            merge_files(main_path=main_p, wood_path=None, gold_path=None,
                        sku_prefix="T", mode="old_parent")


# ===== square 合并模式 =====

class TestSquareMergeMode:
    """合并模式支持正方形 (square) 画作。

    square 检测: main 文件 Size 列预填 L==W (如 12x12) → ratio_type="square"
    square 行为: Product Name 用 SIZES_SQUARE, Length/Width 用正方形尺寸序列
    """

    def test_square_merge_uses_square_sizes(self, tmp_path):
        """square 画作合并: Product Name 和 Length/Width 用正方形尺寸。"""
        from amazon_excel_processor.merger import merge_files
        main_wb, col_map = _create_main_workbook(["Art A"])
        main_ws = main_wb.active
        # 在 main 文件 Size 列预填正方形值, 触发 square 检测
        size_col = col_map["Size"]  # 56
        square_sizes = ["12x12", "16x16", "20x20", "24x24", "28x28"]
        # main 的 10 个子体 (row 9-18)
        for i in range(10):
            main_ws.cell(row=DATA_START_ROW + 1 + i, column=size_col).value = square_sizes[i % 5]

        wood_wb = _create_variant_workbook(["Art A"], role="wood")
        main_p = tmp_path / "main.xlsx"
        wood_p = tmp_path / "wood.xlsx"
        main_wb.save(str(main_p))
        wood_wb.save(str(wood_p))

        out = merge_files(main_path=main_p, wood_path=wood_p, gold_path=None,
                          sku_prefix="T", mode="new")

        # 重新加载验证 (16 行: parent + Frame×5 + Unframe×5 + Wood×5)
        wb = load_workbook(str(out))
        ws = wb["Template"]
        # row 9 = Frame 第1个: Product Name 应含 12x12inch (square 尺寸)
        assert "12x12inch" in str(ws.cell(row=9, column=7).value)
        # row 10 = Frame 第2个: 16x16inch
        assert "16x16inch" in str(ws.cell(row=10, column=7).value)
        # Length 列 (col 124) = 正方形值 12, 16, 20, 24, 28
        assert ws.cell(row=9, column=124).value == 12
        assert ws.cell(row=10, column=124).value == 16
        # Width 列 (col 126) = 正方形值 (L==W)
        assert ws.cell(row=9, column=126).value == 12
        assert ws.cell(row=10, column=126).value == 16
        # Wood 行 (row 19-23) 也用正方形尺寸
        assert "12x12inch" in str(ws.cell(row=19, column=7).value)
        assert ws.cell(row=19, column=124).value == 12
        assert ws.cell(row=19, column=126).value == 12

    def test_32_merge_uses_32_sizes(self, tmp_path):
        """3:2 画作合并 (回归): Product Name 和 Length/Width 用 3:2 尺寸。"""
        from amazon_excel_processor.merger import merge_files
        main_wb, col_map = _create_main_workbook(["Art A"])
        main_ws = main_wb.active
        # Size 列预填 3:2 值 (L != W)
        size_col = col_map["Size"]
        size_32 = ["12x08", "18x12", "24x16", "30x20", "36x24"]
        for i in range(10):
            main_ws.cell(row=DATA_START_ROW + 1 + i, column=size_col).value = size_32[i % 5]

        wood_wb = _create_variant_workbook(["Art A"], role="wood")
        main_p = tmp_path / "main.xlsx"
        wood_p = tmp_path / "wood.xlsx"
        main_wb.save(str(main_p))
        wood_wb.save(str(wood_p))

        out = merge_files(main_path=main_p, wood_path=wood_p, gold_path=None,
                          sku_prefix="T", mode="new")

        wb = load_workbook(str(out))
        ws = wb["Template"]
        # row 9 = Frame 第1个: 3:2 尺寸 08x12inch
        assert "08x12inch" in str(ws.cell(row=9, column=7).value)
        # Length = 12 (3:2 的 _STYLE_LENGTH[0])
        assert ws.cell(row=9, column=124).value == 12
        # Width = 8 (3:2 的 _STYLE_WIDTH[0])
        assert ws.cell(row=9, column=126).value == 8

    def test_mixed_square_and_32_merge(self, tmp_path):
        """混合: 一幅 square 一幅 3:2, 各自用对应尺寸。"""
        from amazon_excel_processor.merger import merge_files
        main_wb, col_map = _create_main_workbook(["Square Art", "Wide Art"])
        main_ws = main_wb.active
        size_col = col_map["Size"]
        # 第1幅 (Square Art, row 8-18): square
        square_sizes = ["12x12", "16x16", "20x20", "24x24", "28x28"]
        for i in range(10):
            main_ws.cell(row=9 + i, column=size_col).value = square_sizes[i % 5]
        # 第2幅 (Wide Art, row 19-29): 3:2
        size_32 = ["12x08", "18x12", "24x16", "30x20", "36x24"]
        for i in range(10):
            main_ws.cell(row=20 + i, column=size_col).value = size_32[i % 5]

        wood_wb = _create_variant_workbook(["Square Art", "Wide Art"], role="wood")
        main_p = tmp_path / "main.xlsx"
        wood_p = tmp_path / "wood.xlsx"
        main_wb.save(str(main_p))
        wood_wb.save(str(wood_p))

        out = merge_files(main_path=main_p, wood_path=wood_p, gold_path=None,
                          sku_prefix="T", mode="new")

        wb = load_workbook(str(out))
        ws = wb["Template"]
        # 第1幅 (16行, row 8-23): square
        assert "12x12inch" in str(ws.cell(row=9, column=7).value)
        assert ws.cell(row=9, column=124).value == 12  # square Length
        assert ws.cell(row=9, column=126).value == 12  # square Width
        # 第2幅 (16行, row 24-39): 3:2
        assert "08x12inch" in str(ws.cell(row=25, column=7).value)
        assert ws.cell(row=25, column=124).value == 12  # 3:2 Length
        assert ws.cell(row=25, column=126).value == 8   # 3:2 Width


# ===== cleanup_for_upload (亚马逊上传前清理) =====

class TestCleanupForUpload:
    """验证 save_workbook 自动清理会导致亚马逊上传失败的字段。

    基于真实文件对比:
      - 成功文件 XL817塔罗杂普: Package Contains 全空, Parent 行包装尺寸全空
      - 失败文件 ZJM817旅游普: Package Contains 全填 1, Parent 行包装尺寸填 1
        → 导致 8007 (父体创建失败) + 990100 (package_contains 关系未批准) + 13013 (子体找不到父体)
    """

    def _make_wb_with_package_fields(self):
        """构造含 Package 字段的 workbook: 1 parent + 2 children。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        headers = {
            1: "SKU", 4: "Parentage Level",
            13: "Package Level",
            14: "Package Contains SKU Quantity",
            15: "Package Contains SKU Identifier",
            208: "Item Package Length", 209: "Package Length Unit",
            210: "Item Package Width", 211: "Package Width Unit",
            212: "Item Package Height", 213: "Package Height Unit",
            214: "Package Weight", 215: "Package Weight Unit",
        }
        for c, h in headers.items():
            ws.cell(row=4, column=c).value = h
        # parent (row 8)
        ws.cell(row=8, column=1).value = "P-1"
        ws.cell(row=8, column=4).value = "Parent"
        ws.cell(row=8, column=13).value = "Unit"
        ws.cell(row=8, column=14).value = 1  # 失败文件: 填了
        ws.cell(row=8, column=15).value = 1
        ws.cell(row=8, column=208).value = 1
        ws.cell(row=8, column=209).value = "Centimeters"
        ws.cell(row=8, column=210).value = 1
        ws.cell(row=8, column=211).value = "Centimeters"
        ws.cell(row=8, column=212).value = 1
        ws.cell(row=8, column=213).value = "Centimeters"
        ws.cell(row=8, column=214).value = 1
        ws.cell(row=8, column=215).value = "Kilograms"
        # child 1 (row 9)
        ws.cell(row=9, column=1).value = "P-1-1"
        ws.cell(row=9, column=4).value = "Child"
        ws.cell(row=9, column=13).value = "Unit"
        ws.cell(row=9, column=14).value = 1  # 失败文件: 填了
        ws.cell(row=9, column=15).value = 1
        ws.cell(row=9, column=208).value = 30  # 子体包装尺寸应保留
        ws.cell(row=9, column=209).value = "Centimeters"
        ws.cell(row=9, column=210).value = 20
        ws.cell(row=9, column=211).value = "Centimeters"
        ws.cell(row=9, column=212).value = 1
        ws.cell(row=9, column=213).value = "Centimeters"
        ws.cell(row=9, column=214).value = 0.18
        ws.cell(row=9, column=215).value = "Kilograms"
        # child 2 (row 10)
        ws.cell(row=10, column=1).value = "P-1-2"
        ws.cell(row=10, column=4).value = "Child"
        ws.cell(row=10, column=13).value = "Unit"
        ws.cell(row=10, column=14).value = 1
        ws.cell(row=10, column=15).value = 1
        ws.cell(row=10, column=208).value = 45
        ws.cell(row=10, column=209).value = "Centimeters"
        return wb

    def test_package_contains_cleared_on_all_rows(self):
        """Package Contains SKU Quantity/Identifier 在所有行 (Parent+Child) 都被清空。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = self._make_wb_with_package_fields()
        ws = wb["Template"]
        # 清理前: 都有值
        assert ws.cell(row=8, column=14).value == 1  # parent
        assert ws.cell(row=9, column=14).value == 1  # child
        cleanup_for_upload(ws)
        # 清理后: 全部为 None
        assert ws.cell(row=8, column=14).value is None  # parent
        assert ws.cell(row=8, column=15).value is None
        assert ws.cell(row=9, column=14).value is None  # child
        assert ws.cell(row=9, column=15).value is None
        assert ws.cell(row=10, column=14).value is None
        assert ws.cell(row=10, column=15).value is None

    def test_parent_package_dimensions_cleared(self):
        """Parent 行的包装尺寸 (Length/Width/Height/Weight + Unit) 被清空。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = self._make_wb_with_package_fields()
        ws = wb["Template"]
        cleanup_for_upload(ws)
        # parent (row 8): 包装尺寸全空
        for c in [208, 209, 210, 211, 212, 213, 214, 215]:
            assert ws.cell(row=8, column=c).value is None, f"col{c} 应被清空"

    def test_child_package_dimensions_preserved(self):
        """Child 行的包装尺寸保留 (子体有实际尺寸)。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = self._make_wb_with_package_fields()
        ws = wb["Template"]
        cleanup_for_upload(ws)
        # child 1 (row 9): 包装尺寸保留
        assert ws.cell(row=9, column=208).value == 30
        assert ws.cell(row=9, column=209).value == "Centimeters"
        assert ws.cell(row=9, column=210).value == 20
        assert ws.cell(row=9, column=214).value == 0.18
        assert ws.cell(row=9, column=215).value == "Kilograms"

    def test_package_level_preserved(self):
        """Package Level (col13) 不被清理, 保持 'Unit'。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = self._make_wb_with_package_fields()
        ws = wb["Template"]
        cleanup_for_upload(ws)
        assert ws.cell(row=8, column=13).value == "Unit"
        assert ws.cell(row=9, column=13).value == "Unit"

    def test_save_workbook_invokes_cleanup(self, tmp_path):
        """save_workbook 保存时自动调用清理, 输出文件中 Package Contains 已清空。"""
        wb = self._make_wb_with_package_fields()
        ws = wb["Template"]
        from amazon_excel_processor.excel_io import save_workbook
        out = save_workbook(ws, tmp_path / "input.xlsx", "Template")
        wb2 = load_workbook(str(out))
        ws2 = wb2["Template"]
        # Package Contains 已清空
        assert ws2.cell(row=8, column=14).value is None
        assert ws2.cell(row=9, column=14).value is None
        # Parent 包装尺寸已清空
        assert ws2.cell(row=8, column=208).value is None
        # Child 包装尺寸保留
        assert ws2.cell(row=9, column=208).value == 30

    def test_no_package_columns_no_error(self):
        """模板没有 Package 列时不报错 (静默跳过)。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        ws.cell(row=4, column=1).value = "SKU"
        ws.cell(row=4, column=4).value = "Parentage Level"
        ws.cell(row=8, column=1).value = "P-1"
        ws.cell(row=8, column=4).value = "Parent"
        # 不应抛异常
        cleanup_for_upload(ws)


# ===== 重复 Item Weight Unit 列清理 =====

class TestDuplicateWeightUnitCleanup:
    """验证双同名列 "Item Weight Unit" 的处理。

    新格式模板中该列名出现 2 次:
      - col148: Item Weight (col147) 右侧, item_weight.unit → 有效列, 保留并填充 Grams
      - col150: Value (col149) 旁, normalized → 无效列, 清空
    """

    def _make_wb_with_duplicate_weight_units(self):
        """构造双 Item Weight Unit 列的 workbook (parent + 2 children, 两列都带值)。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        headers = {
            4: "Parentage Level",
            7: "Item Name",
            147: "Item Weight", 148: "Item Weight Unit",
            149: "Value", 150: "Item Weight Unit",
        }
        for c, h in headers.items():
            ws.cell(row=4, column=c).value = h
        for r, par in [(8, "Parent"), (9, "Child"), (10, "Child")]:
            ws.cell(row=r, column=4).value = par
            ws.cell(row=r, column=147).value = 300
            ws.cell(row=r, column=148).value = "Grams"  # 有效列 (Item Weight 右侧)
            ws.cell(row=r, column=150).value = "Lbs"    # 无效列 (Value 旁) → 应清空
        return wb

    def test_duplicate_weight_unit_cleared_adjacent_kept(self):
        """Value 旁的 Item Weight Unit (col150) 清空, Item Weight 右侧 (col148) 保留。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = self._make_wb_with_duplicate_weight_units()
        ws = wb["Template"]
        cleanup_for_upload(ws)
        for r in [8, 9, 10]:
            assert ws.cell(row=r, column=148).value == "Grams", \
                f"r{r} col148 (Item Weight 右侧) 不应被清空"
            assert ws.cell(row=r, column=150).value is None, \
                f"r{r} col150 (Value 旁) 应被清空"

    def test_single_weight_unit_not_cleared(self):
        """只有一列 Item Weight Unit (无重复) → 不清空。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = self._make_wb_with_duplicate_weight_units()
        ws = wb["Template"]
        ws.cell(row=4, column=150).value = None  # 去掉重复列 → 只剩 col148
        cleanup_for_upload(ws)
        for r in [8, 9, 10]:
            assert ws.cell(row=r, column=148).value == "Grams"

    def test_no_item_weight_col_no_clear(self):
        """没有 Item Weight 列时无法确认有效列 → 不清任何 Item Weight Unit。"""
        from amazon_excel_processor.excel_io import cleanup_for_upload
        wb = self._make_wb_with_duplicate_weight_units()
        ws = wb["Template"]
        ws.cell(row=4, column=147).value = "Other Field"  # Item Weight 列不存在
        cleanup_for_upload(ws)
        # 无法确认哪列有效, 两列都保留 (不误清)
        for r in [8, 9, 10]:
            assert ws.cell(row=r, column=148).value == "Grams"
            assert ws.cell(row=r, column=150).value == "Lbs"

    def test_locate_columns_picks_first_weight_unit(self):
        """locate_columns 对重复列名保留第一个匹配 (Item Weight 右侧 col148)。

        锁定填充目标: fill 逻辑通过 col_map 只写 col148, 不写 col150。
        """
        from amazon_excel_processor.excel_io import locate_columns
        wb = self._make_wb_with_duplicate_weight_units()
        ws = wb["Template"]
        col_map = locate_columns(ws)
        assert col_map["Item Weight"] == 147
        assert col_map["Item Weight Unit"] == 148  # 第一个匹配, 不是 col150

    def test_save_workbook_clears_duplicate_unit(self, tmp_path):
        """save_workbook 保存时自动清空 Value 旁的重复 Item Weight Unit。"""
        from amazon_excel_processor.excel_io import save_workbook
        wb = self._make_wb_with_duplicate_weight_units()
        ws = wb["Template"]
        out = save_workbook(ws, tmp_path / "input.xlsx", "Template")
        wb2 = load_workbook(str(out))
        ws2 = wb2["Template"]
        for r in [8, 9, 10]:
            assert ws2.cell(row=r, column=148).value == "Grams"
            assert ws2.cell(row=r, column=150).value is None


# ===== 合并模式: 源文件带入的重复 Item Weight Unit 值被清空 =====

class TestMergeClearsDuplicateWeightUnit:
    """合并时源文件 (普/木/金) col150 带值 → 输出被清空, col148 填 Grams。"""

    def test_merge_clears_value_adjacent_weight_unit(self, tmp_path):
        """端到端: 源文件 Value 旁 Item Weight Unit (col150) 带脏值 → 输出清空。"""
        from amazon_excel_processor.merger import merge_files
        main_wb, _ = _create_main_workbook(["Art A"])
        main_ws = main_wb.active
        # 补充双 Item Weight Unit 列头 + 源文件 col150 脏值
        main_ws.cell(row=HEADER_ROW, column=147).value = "Item Weight"
        main_ws.cell(row=HEADER_ROW, column=148).value = "Item Weight Unit"
        main_ws.cell(row=HEADER_ROW, column=149).value = "Value"
        main_ws.cell(row=HEADER_ROW, column=150).value = "Item Weight Unit"
        for r in range(DATA_START_ROW, DATA_START_ROW + 11):
            main_ws.cell(row=r, column=150).value = "Lbs"

        wood_wb = _create_variant_workbook(["Art A"], role="wood")
        wood_ws = wood_wb.active
        wood_ws.cell(row=HEADER_ROW, column=147).value = "Item Weight"
        wood_ws.cell(row=HEADER_ROW, column=148).value = "Item Weight Unit"
        wood_ws.cell(row=HEADER_ROW, column=149).value = "Value"
        wood_ws.cell(row=HEADER_ROW, column=150).value = "Item Weight Unit"
        for r in range(DATA_START_ROW, DATA_START_ROW + 6):
            wood_ws.cell(row=r, column=150).value = "Lbs"

        main_p = tmp_path / "main.xlsx"
        wood_p = tmp_path / "wood.xlsx"
        main_wb.save(str(main_p))
        wood_wb.save(str(wood_p))

        out = merge_files(main_path=main_p, wood_path=wood_p, gold_path=None,
                          sku_prefix="T", mode="new")

        ws = load_workbook(str(out))["Template"]
        # 全部 16 行: col150 (Value 旁) 空, col148 (Item Weight 右侧) = Grams
        for r in range(DATA_START_ROW, DATA_START_ROW + 16):
            assert ws.cell(row=r, column=148).value == "Grams", \
                f"r{r} col148 应填 Grams"
            assert ws.cell(row=r, column=150).value is None, \
                f"r{r} col150 (Value 旁) 应为空"

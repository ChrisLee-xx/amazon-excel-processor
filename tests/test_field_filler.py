"""变体字段填充模块测试 — detect_ratio_type

旧的单字段填充函数 (fill_color/fill_size/fill_length 等) 已移除,
新格式统一使用 fill_group_merged (在 test_merger.py 中测试)。
本文件仅保留 detect_ratio_type 的单元测试。
"""

from openpyxl import Workbook

from amazon_excel_processor.field_filler import detect_ratio_type


def _create_test_ws(product_names: list[str]) -> tuple:
    """创建测试用 worksheet，返回 (ws, rows, col_map)。"""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Item Name"
    ws.cell(row=1, column=2).value = "Color"
    ws.cell(row=1, column=3).value = "Size"

    rows = []
    for i, name in enumerate(product_names):
        row = i + 2
        ws.cell(row=row, column=1).value = name
        rows.append(row)

    col_map = {"Item Name": 1, "Color": 2, "Size": 3}
    return ws, rows, col_map


def _make_32_names():
    """生成 3:2 比例的 11 行产品名称。"""
    return [
        "Parent Title",
        f'Title Frame-style 12"L x 8"W',
        f'Title Frame-style 18"L x 12"W',
        f'Title Frame-style 24"L x 16"W',
        f'Title Frame-style 30"L x 20"W',
        f'Title Frame-style 36"L x 24"W',
        f'Title Unframe-style 12"L x 8"W',
        f'Title Unframe-style 18"L x 12"W',
        f'Title Unframe-style 24"L x 16"W',
        f'Title Unframe-style 30"L x 20"W',
        f'Title Unframe-style 36"L x 24"W',
    ]


def _make_square_names():
    """生成正方形比例的 11 行产品名称。"""
    return [
        "Parent Title",
        f'Title Frame-style 12"L x 12"W',
        f'Title Frame-style 16"L x 16"W',
        f'Title Frame-style 20"L x 20"W',
        f'Title Frame-style 24"L x 24"W',
        f'Title Frame-style 28"L x 28"W',
        f'Title Unframe-style 12"L x 12"W',
        f'Title Unframe-style 16"L x 16"W',
        f'Title Unframe-style 20"L x 20"W',
        f'Title Unframe-style 24"L x 24"W',
        f'Title Unframe-style 28"L x 28"W',
    ]


class TestDetectRatioType:
    def test_32_ratio_size_empty(self):
        """Size 列为空 → 3:2"""
        ws, rows, col_map = _create_test_ws(_make_32_names())
        assert detect_ratio_type(ws, rows, col_map) == "3:2"

    def test_32_ratio_size_prefilled_unequal(self):
        """Size 列预填值 L!=W（如 12"L x 8"W）→ 3:2"""
        ws, rows, col_map = _create_test_ws(_make_32_names())
        size_col = col_map["Size"]
        # 3:2 尺寸: L != W
        size_values = ["", "12x08", "18x12", "24x16", "30x20", "36x24",
                       "12x08", "18x12", "24x16", "30x20", "36x24"]
        for i, row in enumerate(rows):
            ws.cell(row=row, column=size_col).value = size_values[i]
        assert detect_ratio_type(ws, rows, col_map) == "3:2"

    def test_square_ratio_size_prefilled(self):
        """Size 列预填 L==W（如 12x12）→ square"""
        ws, rows, col_map = _create_test_ws(_make_square_names())
        size_col = col_map["Size"]
        # 正方形尺寸: L == W
        size_values = ["", "12x12", "16x16", "20x20", "24x24", "28x28",
                       "12x12", "16x16", "20x20", "24x24", "28x28"]
        for i, row in enumerate(rows):
            ws.cell(row=row, column=size_col).value = size_values[i]
        assert detect_ratio_type(ws, rows, col_map) == "square"

    def test_no_size_column_defaults_32(self):
        """没有 Size 列 → 默认 3:2"""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Item Name"
        ws.cell(row=2, column=1).value = "Title"
        col_map = {"Item Name": 1}  # 无 Size
        assert detect_ratio_type(ws, [2], col_map) == "3:2"

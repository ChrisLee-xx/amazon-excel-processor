"""Excel 文件读写模块 (新格式)

亚马逊新模板格式:
- Row 4 = 列名 (header)
- Row 8 = 数据起始行
- 关键列: SKU(col1), Parentage Level(col4), Parent SKU(col5),
  Variation Theme Name(col6), Item Name(col7), Style(col46), Color(col55),
  Size(col56), Item Length Longer Edge(col124), Item Width Shorter Edge(col126),
  Item Weight(col147), List Price(col154)
"""

import logging
import os
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook as _load_wb
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = ["Item Name"]

OPTIONAL_COLUMNS = [
    "SKU",
    "Parent SKU",
    "Parentage Level",
    "Variation Theme Name",
    "Style",
    "Color",
    "Size",
    "Item Length Longer Edge",
    "Item Length Unit",
    "Item Width Shorter Edge",
    "Item Width Unit",
    "Item Weight",
    "Item Weight Unit",
    "List Price",
    # Shipping (Package) 列
    "Item Package Length",
    "Package Length Unit",
    "Item Package Width",
    "Package Width Unit",
    "Item Package Height",
    "Package Height Unit",
    "Package Weight",
    "Package Weight Unit",
]

HEADER_ROW = 4      # 新格式: 列名在第 4 行
DATA_START_ROW = 8  # 新格式: 数据从第 8 行开始
GROUP_SIZE = 11     # 单文件模式 11 行/组; 合并时木/金用 6, 输出用 21


def load_workbook(filepath: str | Path):
    """读取 Excel 文件，只保留 Template sheet 以加速处理。"""
    filepath = Path(filepath)

    if filepath.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"不支持的文件格式: {filepath.suffix}，仅支持 .xlsx 和 .xlsm")

    keep_vba = filepath.suffix.lower() == ".xlsm"
    wb = _load_wb(str(filepath), keep_vba=keep_vba)
    logger.debug("openpyxl 加载完成, sheets=%s", wb.sheetnames)

    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == "template":
            sheet_name = name
            break

    if sheet_name is None:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"找不到 'template' sheet。可用的 sheet: {available}")

    for name in list(wb.sheetnames):
        if name != sheet_name:
            del wb[name]

    ws = wb[sheet_name]
    logger.debug("Template sheet: max_row=%d, max_column=%d", ws.max_row, ws.max_column)
    return wb, ws, sheet_name


def locate_columns(ws: Worksheet, header_row: int = HEADER_ROW) -> dict[str, int]:
    """扫描表头行 (第 4 行) 动态定位列索引。

    新格式模板中 Row 4 是列名, Row 8 是数据。
    返回 {列名: 列号(1-based)} 的映射。
    """
    col_map: dict[str, int] = {}

    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value is None:
            continue
        header = str(cell_value).strip()
        all_columns = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
        for expected in all_columns:
            if header.lower() == expected.lower():
                col_map[expected] = col_idx
                break

    for req in REQUIRED_COLUMNS:
        if req not in col_map:
            raise ValueError(f"必需列 '{req}' 在表头中未找到 (新格式应位于第 4 行)")

    found_optional = [c for c in OPTIONAL_COLUMNS if c in col_map]
    missing_optional = [c for c in OPTIONAL_COLUMNS if c not in col_map]
    if missing_optional:
        logger.info("可选列未找到（将跳过）: %s", ", ".join(missing_optional))
    if found_optional:
        logger.info("已定位列: %s", ", ".join([*REQUIRED_COLUMNS, *found_optional]))
    logger.debug("locate_columns: col_map=%s", col_map)

    return col_map


def _find_last_data_row(ws: Worksheet, start_row: int = DATA_START_ROW) -> int:
    """找到最后一个真正数据行的行号。

    数据行的判定: Parentage Level 列 (col4) 有值 (Parent/Child)。
    表格底部的备注/说明行 (如 "SKU命名（一样）"、"普/木/金" 等手动注释)
    通常 Parentage Level 为空, 会被正确排除。

    向下多探测 20 行, 防止 max_row 偏小 (Windows 平台差异)。
    """
    check_limit = ws.max_row + 20
    last_data_row = start_row - 1
    logger.debug("_find_last_data_row: ws.max_row=%d, check_limit=%d", ws.max_row, check_limit)

    for row in range(start_row, check_limit + 1):
        # 用 Parentage Level (col4) 判断是否是数据行
        parentage = ws.cell(row=row, column=4).value
        has_data = parentage is not None and str(parentage).strip() != ""
        if has_data:
            last_data_row = row

    logger.debug("_find_last_data_row: result=%d (数据行数=%d)",
                 last_data_row, last_data_row - start_row + 1 if last_data_row >= start_row else 0)
    return last_data_row


def group_rows(ws: Worksheet, group_size: int = GROUP_SIZE) -> list[list[int]]:
    """将数据行按 group_size 行一组分组。

    返回 [[row_num, ...], ...] 列表。
    不完整尾部组记录警告并跳过。

    Args:
        ws: 目标 worksheet
        group_size: 每组行数; 默认 11 (单文件模式),
                    合并时木/金用 6, 输出用 21
    """
    last_row = _find_last_data_row(ws)
    data_rows = list(range(DATA_START_ROW, last_row + 1))
    logger.debug("group_rows: last_data_row=%d, total_data_rows=%d, group_size=%d",
                 last_row, len(data_rows), group_size)

    if not data_rows:
        logger.warning("template sheet 没有数据行")
        return []

    total = len(data_rows)
    complete_groups = total // group_size
    remainder = total % group_size

    if remainder > 0:
        logger.warning(
            "数据行数 %d 不是 %d 的倍数，尾部 %d 行将被跳过",
            total, group_size, remainder,
        )

    groups = []
    for i in range(complete_groups):
        start = i * group_size
        group = data_rows[start: start + group_size]
        groups.append(group)

    logger.debug("group_rows: %d 个完整组, group_size=%d, 首组=%s, 末组=%s",
                 len(groups), group_size,
                 groups[0] if groups else "N/A",
                 groups[-1] if groups else "N/A")
    return groups


def _can_write(path: Path) -> bool:
    """检测文件是否可写入（未被其他进程锁定）。"""
    if not path.exists():
        return True
    try:
        os.remove(str(path))
        return True
    except (PermissionError, OSError):
        return False


def _resolve_output_path(input_path: Path, output_path: Optional[Path], suffix: str = "_processed") -> Path:
    """确定输出文件路径，若目标被占用则自动加序号。"""
    if output_path is None:
        base = input_path.parent / f"{input_path.stem}{suffix}{input_path.suffix}"
    else:
        base = output_path

    if _can_write(base):
        logger.debug("_resolve_output_path: 使用 %s", base.name)
        return base

    for i in range(2, 100):
        candidate = base.parent / f"{base.stem}_{i}{base.suffix}"
        if _can_write(candidate):
            logger.warning("输出文件 %s 被占用，改用 %s", base.name, candidate.name)
            return candidate

    raise OSError(f"无法创建输出文件: {base}")


def save_workbook(
    ws: Worksheet,
    input_path: str | Path,
    template_name: str,
    output_path: Optional[str | Path] = None,
    suffix: str = "_processed",
):
    """保存 worksheet 为新的 Excel 文件，保留 VBA 宏。

    保存前自动调用 cleanup_for_upload 清理会导致亚马逊上传失败的字段。
    """
    cleanup_for_upload(ws)
    input_path = Path(input_path)
    wb = ws.parent
    out = _resolve_output_path(input_path, Path(output_path) if output_path else None, suffix=suffix)
    wb.save(str(out))
    logger.info("保存文件: %s", out)
    return out


# 亚马逊上传时会导致错误的字段 (需在保存前清空)
# Package Contains 字段仅用于套装商品, 普通单品填了会导致 8007/990100 错误
_CLEANUP_COLUMNS_BOTH = [
    "Package Contains SKU Quantity",
    "Package Contains SKU Identifier",
]
# Parent 行 (虚拟父体) 不应有实际包装尺寸, 只有 Child 行才保留
_CLEANUP_COLUMNS_PARENT_ONLY = [
    "Item Package Length",
    "Package Length Unit",
    "Item Package Width",
    "Package Width Unit",
    "Item Package Height",
    "Package Height Unit",
    "Package Weight",
    "Package Weight Unit",
]


def _find_col_by_header(ws: Worksheet, header_name: str, header_row: int = HEADER_ROW) -> int:
    """按列名查找列号 (不依赖 col_map), 找不到返回 0。"""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip().lower() == header_name.lower():
            return c
    return 0


def cleanup_for_upload(ws: Worksheet) -> None:
    """清理会导致亚马逊上传失败的字段。

    规则 (基于成功上传文件 XL817塔罗杂普_processed 与失败文件 ZJM817旅游普_processed 的对比):
      1. Package Contains SKU Quantity / Identifier (col14/15):
         全部行清空 (Parent + Child)。这两个字段仅用于套装商品,
         普通单品填了会导致 990100 警告 + 8007 父体创建失败。
      2. Parent 行的 Item Package Length/Width/Height/Weight + 对应 Unit (col208-215):
         全部 Parent 行清空。虚拟父体没有实际包装尺寸, 只有 Child 子体才保留。
    """
    # 1. 全部行: 清空 Package Contains 字段
    cols_both = [_find_col_by_header(ws, n) for n in _CLEANUP_COLUMNS_BOTH]
    cols_both = [c for c in cols_both if c > 0]

    # 2. 仅 Parent 行: 清空包装尺寸字段
    cols_parent = [_find_col_by_header(ws, n) for n in _CLEANUP_COLUMNS_PARENT_ONLY]
    cols_parent = [c for c in cols_parent if c > 0]

    if not cols_both and not cols_parent:
        return  # 模板里没这些列, 无需清理

    parentage_col = _find_col_by_header(ws, "Parentage Level")
    if parentage_col == 0:
        parentage_col = 4  # 默认列号

    cleaned_both = 0
    cleaned_parent = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        parentage = ws.cell(row=r, column=parentage_col).value
        if parentage is None or str(parentage).strip() == "":
            continue  # 非数据行

        # 清空 Package Contains (所有数据行)
        for c in cols_both:
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None
                cleaned_both += 1

        # 清空 Parent 行的包装尺寸
        if str(parentage).strip().lower() == "parent":
            for c in cols_parent:
                if ws.cell(row=r, column=c).value is not None:
                    ws.cell(row=r, column=c).value = None
                    cleaned_parent += 1

    if cleaned_both or cleaned_parent:
        logger.info("cleanup_for_upload: 清空 Package Contains %d 格, Parent 包装尺寸 %d 格",
                    cleaned_both, cleaned_parent)


def copy_cell_style(src_cell, dst_cell) -> None:
    """直接复制源单元格的内部 _style 数组到目标单元格.

    比 openpyxl 的 `copy(src)` 更稳定: 不会清空 dst 的 value,
    只替换 fill/font/border/alignment/number_format 等视觉样式。

    用于把模板单元格的"涂黑/涂色"格式精确复制到目标位置
    (如把 E8 的深色填充复制到新生成 group 的 parent 行 Parent SKU 列)。
    """
    from copy import copy as _copy
    dst_cell._style = _copy(src_cell._style)

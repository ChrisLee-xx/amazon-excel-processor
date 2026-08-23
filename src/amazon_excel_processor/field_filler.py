"""变体字段填充模块"""

import logging
import re

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ===== 合并模式 21 元素序列 =====
# 结构: [parent, Frame×5, Unframe×5, VintageWood×5, VintageOrnate×5]
# Wood 和 Gold 的 Size/SizeMap/Length/Width/Weight 与 Frame×5 一致
# Wood 和 Gold 的 Price 为: 26.9 / 39.9 / 59.9 / 99.9 / 129.9

COLOR_SEQUENCE_21 = [
    "",
    "Frame-style", "Frame-style", "Frame-style", "Frame-style", "Frame-style",
    "Unframe-style", "Unframe-style", "Unframe-style", "Unframe-style", "Unframe-style",
    "Vintage Wood Grain Frame-style", "Vintage Wood Grain Frame-style",
    "Vintage Wood Grain Frame-style", "Vintage Wood Grain Frame-style", "Vintage Wood Grain Frame-style",
    "Vintage Ornate Gold Frame-style", "Vintage Ornate Gold Frame-style",
    "Vintage Ornate Gold Frame-style", "Vintage Ornate Gold Frame-style", "Vintage Ornate Gold Frame-style",
]

SIZE_MAP_SEQUENCE_21 = [
    "",
    "X-Small", "Small", "Medium", "Large", "X-Large",
    "X-Small", "Small", "Medium", "Large", "X-Large",
    "X-Small", "Small", "Medium", "Large", "X-Large",
    "X-Small", "Small", "Medium", "Large", "X-Large",
]

SIZE_32_21 = [
    "",
    "12L''x08W''", "18L''x12W''", "24L''x16W''", "30L''x20W''", "36L''x24W''",
    "12L''x08W''", "18L''x12W''", "24L''x16W''", "30L''x20W''", "36L''x24W''",
    "12L''x08W''", "18L''x12W''", "24L''x16W''", "30L''x20W''", "36L''x24W''",
    "12L''x08W''", "18L''x12W''", "24L''x16W''", "30L''x20W''", "36L''x24W''",
]

LENGTH_32_21 = [""] + [12, 18, 24, 30, 36] * 4  # parent + 4 styles × 5 sizes = 21 (英寸)

WIDTH_32_21 = [""] + [8, 12, 16, 20, 24] * 4

# Weight: parent=空, Frame 和 Wood/Gold 均为 0.18-0.88, Unframe 为 0.02-0.25
WEIGHT_SEQUENCE_21 = [
    1,
    300, 400, 600, 1000, 1500,   # Frame×5
    80, 90, 130, 180, 240,       # Unframe×5
    450, 850, 1500, 2400, 3400,  # Wood×5
    450, 850, 1500, 2400, 3400,  # Gold×5
]

# Shipping (Package) — 21 行: parent 空, 普通 10 行, 木金 10 行
# 普通 Frame×5 + Unframe×5: L/W = (30,20)(45,30)(60,40)(75,50)(90,60) 各 2 段
# 木金 Wood×5 + Gold×5:    L/W = (32,22)(47,32)(62,42)(77,52)(92,62)
# Height: 1 (普通) / 4.5 (木金)
# Weight: 0.18/0.28/0.48/0.68/0.88 (普通 5 + 木金 5), 0.02/0.04/0.07/0.15/0.25 (普通 unframe 5)
PACKAGE_LENGTH_21 = [
    "",
    30, 45, 60, 75, 90,           # Frame×5
    30, 45, 60, 75, 90,           # Unframe×5
    32, 47, 62, 77, 92,           # Wood×5
    32, 47, 62, 77, 92,           # Gold×5
]
PACKAGE_WIDTH_21 = [
    "",
    20, 30, 40, 50, 60,           # Frame×5
    20, 30, 40, 50, 60,           # Unframe×5
    22, 32, 42, 52, 62,           # Wood×5
    22, 32, 42, 52, 62,           # Gold×5
]
PACKAGE_HEIGHT_21 = [
    "",
    1, 1, 1, 1, 1,                # Frame×5
    1, 1, 1, 1, 1,                # Unframe×5
    4.5, 4.5, 4.5, 4.5, 4.5,      # Wood×5
    4.5, 4.5, 4.5, 4.5, 4.5,      # Gold×5
]
PACKAGE_WEIGHT_21 = [
    "",
    0.18, 0.28, 0.48, 0.68, 0.88,   # Frame×5
    0.02, 0.04, 0.07, 0.15, 0.25,   # Unframe×5
    0.18, 0.28, 0.48, 0.68, 0.88,   # Wood×5
    0.18, 0.28, 0.48, 0.68, 0.88,   # Gold×5
]
PACKAGE_LENGTH_UNIT_21 = "Centimeters"  # 全组统一
PACKAGE_WIDTH_UNIT_21 = "Centimeters"
PACKAGE_HEIGHT_UNIT_21 = "Centimeters"
PACKAGE_WEIGHT_UNIT_21 = "Kilograms"


# Price: parent=空, Frame: 19.9/29.9/45/75/99, Unframe: 11.9/14.9/19.9/24.9/34.9,
#        Wood: 26.9/39.9/59.9/99.9/129.9, Gold: 26.9/39.9/59.9/99.9/129.9
PRICE_SEQUENCE_21 = [
    "",
    19.9, 29.9, 45, 75, 99,         # Frame×5
    11.9, 14.9, 19.9, 24.9, 34.9,   # Unframe×5
    26.9, 39.9, 59.9, 99.9, 129.9,  # Wood×5
    26.9, 39.9, 59.9, 99.9, 129.9,  # Gold×5
]

# List Price = Your Price (每行同步填)
# 这里直接指向 PRICE_SEQUENCE_21, merger 阶段可共用

# 新格式模板的 3 个价格列, 全部填同一价格序列
# (价格按 size 从小到大排列, 与 Size 列顺序一致)
PRICE_COLUMNS = [
    "List Price",
    "Your Price USD (Sell on Amazon, US)",
    "Your Price USD (Amazon Business (B2B), US)",
]


# ===== 动态 style 计划 (木/金可选) =====
# Frame+Unframe 永远来自普文件 (固定 11 行); Wood/Gold 按需追加在后面。
# 每个 style 的 5 尺寸字段值从 STYLE_SPECS 拼接出逐行序列再填充。
# 当 active_styles=[frame,unframe,wood,gold] 时, 动态序列与上面 *_21 常量逐元素相等。

# 所有 style 共享的 5 尺寸值 (新格式, 英寸)
_STYLE_SIZE_MAP = ["X-Small", "Small", "Medium", "Large", "X-Large"]
# 3:2 比例
_STYLE_SIZE_32 = ["12L''x08W''", "18L''x12W''", "24L''x16W''", "30L''x20W''", "36L''x24W''"]
_STYLE_LENGTH = [12, 18, 24, 30, 36]
_STYLE_WIDTH = [8, 12, 16, 20, 24]
# 正方形比例 (L == W)
_STYLE_SIZE_SQUARE = ["12L''x12W''", "16L''x16W''", "20L''x20W''", "24L''x24W''", "28L''x28W''"]
_STYLE_LENGTH_SQUARE = [12, 16, 20, 24, 28]
_STYLE_WIDTH_SQUARE = [12, 16, 20, 24, 28]
# edge 序列已废弃: Style 列保留原始值, 不再填充 (Length 列即 Longer Edge)

STYLE_SPECS = {
    "frame": {
        "label": "Frame-style",
        # 新格式: Item Weight 单位为 Grams (克)
        "weight": [300, 400, 600, 1000, 1500],
        "price": [19.9, 29.9, 45, 75, 99],
        # Package (Shipping) — 普通: L/W = (30,20)(45,30)(60,40)(75,50)(90,60)
        "package_length": [30, 45, 60, 75, 90],
        "package_width": [20, 30, 40, 50, 60],
        "package_height": [1, 1, 1, 1, 1],
        "package_weight": [0.18, 0.28, 0.48, 0.68, 0.88],
    },
    "unframe": {
        "label": "Unframe-style",
        "weight": [80, 90, 130, 180, 240],
        "price": [11.9, 14.9, 19.9, 24.9, 34.9],
        # Package — 普通 unframe: L/W 同 frame 序列, 重量更小 (0.02-0.25)
        "package_length": [30, 45, 60, 75, 90],
        "package_width": [20, 30, 40, 50, 60],
        "package_height": [1, 1, 1, 1, 1],
        "package_weight": [0.02, 0.04, 0.07, 0.15, 0.25],
    },
    "wood": {
        "label": "Vintage Wood Grain Frame-style",
        "weight": [450, 850, 1500, 2400, 3400],
        "price": [26.9, 39.9, 59.9, 99.9, 129.9],
        # Package — 木金: L/W = (32,22)(47,32)(62,42)(77,52)(92,62), H=4.5, Weight 大
        "package_length": [32, 47, 62, 77, 92],
        "package_width": [22, 32, 42, 52, 62],
        "package_height": [4.5, 4.5, 4.5, 4.5, 4.5],
        "package_weight": [0.18, 0.28, 0.48, 0.68, 0.88],
    },
    "gold": {
        "label": "Vintage Ornate Gold Frame-style",
        "weight": [450, 850, 1500, 2400, 3400],
        "price": [26.9, 39.9, 59.9, 99.9, 129.9],
        "package_length": [32, 47, 62, 77, 92],
        "package_width": [22, 32, 42, 52, 62],
        "package_height": [4.5, 4.5, 4.5, 4.5, 4.5],
        "package_weight": [0.18, 0.28, 0.48, 0.68, 0.88],
    },
}

# 永远来自普文件的 style (固定前 10 个 child)
MAIN_STYLES = ["frame", "unframe"]
# 可选变体 style (按输出顺序追加)
VARIANT_STYLES = ["wood", "gold"]


def build_active_styles(has_wood: bool, has_gold: bool) -> list:
    """返回合并输出的 style 顺序 (frame, unframe 总在, wood/gold 按需)."""
    styles = list(MAIN_STYLES)
    if has_wood:
        styles.append("wood")
    if has_gold:
        styles.append("gold")
    return styles


def _build_sequences(active_styles: list, ratio_type: str = "3:2") -> dict:
    """根据 active_styles 构建各字段的逐行序列 (含 parent 行占位).

    Args:
        active_styles: 参与的 style 列表 (如 ["frame","unframe","wood"])
        ratio_type: "3:2" 或 "square", 决定 Size/Length/Width 用哪套尺寸

    返回 dict, 每个序列长度 = 1 + 5*len(active_styles):
      color / size_map / size_32 / length / width / weight / price / labels
    parent 行: color="", size_map="", size_32="", length="", width="",
              weight="", price="", labels=None
    """
    # 按比例选择尺寸序列
    if ratio_type == "square":
        size_seq = _STYLE_SIZE_SQUARE
        length_seq = _STYLE_LENGTH_SQUARE
        width_seq = _STYLE_WIDTH_SQUARE
    else:
        size_seq = _STYLE_SIZE_32
        length_seq = _STYLE_LENGTH
        width_seq = _STYLE_WIDTH

    seqs = {
        "color": [""],
        "size_map": [""],
        "size_32": [""],
        "length": [""],
        "width": [""],
        "weight": [1],
        "price": [""],
        "labels": [None],
        # Shipping (Package)
        "package_length": [""],
        "package_width": [""],
        "package_height": [""],
        "package_weight": [""],
    }
    for key in active_styles:
        spec = STYLE_SPECS[key]
        label = spec["label"]
        seqs["color"].extend([label] * 5)
        seqs["size_map"].extend(_STYLE_SIZE_MAP)
        seqs["size_32"].extend(size_seq)
        seqs["length"].extend(length_seq)
        seqs["width"].extend(width_seq)
        seqs["weight"].extend(spec["weight"])
        seqs["price"].extend(spec["price"])
        seqs["labels"].extend([label] * 5)
        seqs["package_length"].extend(spec["package_length"])
        seqs["package_width"].extend(spec["package_width"])
        seqs["package_height"].extend(spec["package_height"])
        seqs["package_weight"].extend(spec["package_weight"])
    return seqs


def fill_group_merged(
    ws: Worksheet,
    rows: list[int],
    col_map: dict[str, int],
    ratio_type: str,
    active_styles: list,
) -> None:
    """编排合并产品组的所有字段填充 (动态 style 数)。

    新格式列映射:
      - Length 列 (Item Length Longer Edge, col124) 填英寸长度 12/18/24/30/36
      - Width 列 (Item Width Shorter Edge, col126) 填英寸宽度 8/12/16/20/24
      - 价格列 (List Price col154 / Your Price col182 / B2B col191) 填同一价格序列
      - Weight 列 (col147) 单位克
      - Style (col46) 填 style 标签
    """
    seqs = _build_sequences(active_styles, ratio_type)

    # 简单字段 (全组相同): 新格式 Variation Theme Name (col6) = "COLOR/SIZE"
    simple_fills = {
        "Variation Theme Name": "COLOR/SIZE",
        "Paint Type": "Oil",
        "Color Map": "Multi",
    }
    for field_name, value in simple_fills.items():
        if field_name not in col_map:
            continue
        col_idx = col_map[field_name]
        for row in rows:
            ws.cell(row=row, column=col_idx).value = value

    # 逐行序列字段
    _fill_seq(ws, rows, col_map, "Color", seqs["color"])
    if ratio_type != "square":
        _fill_seq(ws, rows, col_map, "Size", seqs["size_32"])
    _fill_seq(ws, rows, col_map, "Size Map", seqs["size_map"])
    _fill_seq(ws, rows, col_map, "Item Length Longer Edge", seqs["length"])
    _fill_seq(ws, rows, col_map, "Item Width Shorter Edge", seqs["width"])
    _fill_seq(ws, rows, col_map, "Item Weight", seqs["weight"])
    # 新格式: 3 个价格列 (List Price / Your Price / B2B) 填同一价格序列
    # 价格随 size 从小到大排列, 与 Size 列顺序一致
    for price_col in PRICE_COLUMNS:
        _fill_seq(ws, rows, col_map, price_col, seqs["price"])
    # 注意: Style 列保留原始值, 不覆盖 (Color 列才填 style 标签)

    # Shipping (Package) 字段填充
    _fill_seq(ws, rows, col_map, "Item Package Length", seqs["package_length"])
    _fill_seq(ws, rows, col_map, "Item Package Width", seqs["package_width"])
    _fill_seq(ws, rows, col_map, "Item Package Height", seqs["package_height"])
    _fill_seq(ws, rows, col_map, "Package Weight", seqs["package_weight"])

    # Unit 列全组统一填充
    if "Item Length Unit" in col_map:
        _fill_const(ws, rows, col_map["Item Length Unit"], "Inches")
    if "Item Width Unit" in col_map:
        _fill_const(ws, rows, col_map["Item Width Unit"], "Inches")
    if "Item Weight Unit" in col_map:
        _fill_const(ws, rows, col_map["Item Weight Unit"], "Grams")
    if "Package Length Unit" in col_map:
        _fill_const(ws, rows, col_map["Package Length Unit"], "Centimeters")
    if "Package Width Unit" in col_map:
        _fill_const(ws, rows, col_map["Package Width Unit"], "Centimeters")
    if "Package Height Unit" in col_map:
        _fill_const(ws, rows, col_map["Package Height Unit"], "Centimeters")
    if "Package Weight Unit" in col_map:
        _fill_const(ws, rows, col_map["Package Weight Unit"], "Kilograms")

    # Search Terms: 下划线替换为空格 (与老品模式 _fill_variant_fields 保持一致)
    clean_search_terms(ws, rows, col_map)


def _fill_const(ws, rows, col_idx, value):
    """填充所有行为同一个常量值."""
    for row in rows:
        ws.cell(row=row, column=col_idx).value = value


def _fill_seq(ws, rows, col_map, field_name, sequence):
    """按 sequence 逐行填充某列 (sequence 长度需 >= len(rows))."""
    if field_name not in col_map:
        return
    col_idx = col_map[field_name]
    for i, row in enumerate(rows):
        ws.cell(row=row, column=col_idx).value = sequence[i]


def detect_ratio_type(
    ws: Worksheet,
    rows: list[int],
    col_map: dict[str, int],
) -> str:
    """检测产品组的比例类型。

    解析 Size 列预填值中的两个数字，L==W 为正方形，L!=W 为 3:2。
    Size 列为空时默认 3:2（由脚本后续填充）。
    返回 "square" 或 "3:2"。
    """
    if "Size" not in col_map:
        return "3:2"
    size_col = col_map["Size"]
    for i, row in enumerate(rows):
        if i == 0:  # 跳过 parent 行（本就为空）
            continue
        value = ws.cell(row=row, column=size_col).value
        if value is None or not str(value).strip():
            continue
        # 解析 Size 值中的数字，比较前两个判断长宽是否相等
        numbers = re.findall(r"\d+", str(value))
        if len(numbers) >= 2:
            return "square" if numbers[0] == numbers[1] else "3:2"
        # 格式无法解析但有值，保守判断为正方形
        return "square"
    return "3:2"


def clean_search_terms(
    ws: Worksheet,
    rows: list[int],
    col_map: dict[str, int],
) -> None:
    """将 Search Terms 列中的下划线替换为空格。"""
    if "Search Terms" not in col_map:
        return
    col_idx = col_map["Search Terms"]
    for row in rows:
        value = ws.cell(row=row, column=col_idx).value
        if value is not None and isinstance(value, str) and "_" in value:
            ws.cell(row=row, column=col_idx).value = value.replace("_", " ")


def fill_group(
    ws: Worksheet,
    rows: list[int],
    col_map: dict[str, int],
    ratio_type: str,
) -> None:
    """编排单个产品组 (11 行, Frame+Unframe) 的所有字段填充。

    新格式下等价于 fill_group_merged(active_styles=["frame","unframe"])。
    """
    fill_group_merged(ws, rows, col_map, ratio_type,
                      build_active_styles(has_wood=False, has_gold=False))

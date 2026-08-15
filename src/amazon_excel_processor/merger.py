"""三文件合并模块

输入文件 (木/金可选):
  - main_path: 普文件 (11 行/组, 含 Frame+Unframe 2 个 style), 必填
  - wood_path: 木框文件 (6 行/组, 每画 1 个 group, 对应 Vintage Wood Grain Frame-style), 可选
  - gold_path: 金框文件 (6 行/组, 每画 1 个 group, 对应 Vintage Ornate Gold Frame-style), 可选

合并输出每组行数 = 1 + 5×(2 + 有木 + 有金): 11 / 16 / 21, 顺序固定:
  1. Frame-style (5 尺寸, 来自 main)
  2. Unframe-style (5 尺寸, 来自 main)
  3. Vintage Wood Grain Frame-style (5 尺寸, 来自 wood, 若提供)
  4. Vintage Ornate Gold Frame-style (5 尺寸, 来自 gold, 若提供)

用户在 GUI 中按 [主, 木, 金] 顺序指定文件 (木/金可留空跳过), 不再依赖"第 1 次/第 2 次"假设.
"""

import logging
import re
from pathlib import Path
from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet

from .excel_io import (
    DATA_START_ROW,
    group_rows,
    load_workbook,
    locate_columns,
    save_workbook,
)
from .field_filler import (
    fill_group_merged,
    build_active_styles,
    _build_sequences,
    STYLE_SPECS,
    detect_ratio_type,
)
from .name_normalizer import (
    SIZES_32,
    SIZES_SQUARE,
    extract_base_title,
    remove_numeric_suffix,
    deduplicate_words,
    collapse_spaces,
)

logger = logging.getLogger(__name__)

MERGED_GROUP_SIZE = 21  # 最大: parent + 4 style × 5 size (wood+gold 都在)
VARIANT_GROUP_SIZE = 6   # 木/金文件每画 1 个 group, 6 行
MAIN_GROUP_SIZE = 11     # 普文件每画 1 个 group, 11 行


def merged_group_size(has_wood: bool, has_gold: bool) -> int:
    """合并输出每组行数 = 1 (parent) + 5 × (2 + 有木 + 有金)。

    木/金皆无 → 11; 只有其一 → 16; 都在 → 21。
    """
    return 1 + 5 * (2 + bool(has_wood) + bool(has_gold))

WOOD_STYLE = "Vintage Wood Grain Frame-style"
GOLD_STYLE = "Vintage Ornate Gold Frame-style"

# 新格式 Item Name 的 style 标签 (与 Color 列不同!):
# Wood 的 Item Name 带 "Vintage Ornate Gold" 前缀 (复现最终文件), 但 Color 列是干净的
ITEM_STYLE_LABELS = {
    "frame": "Frame-style",
    "unframe": "Unframe-style",
    "wood": "Vintage Ornate Gold Vintage Wood Grain Frame-style",
    "gold": "Vintage Ornate Gold Frame-style",
}

# 新格式列号 (新格式 header row 4, data row 8)
COL_SELLER_SKU = 1       # SKU
COL_PRODUCT_NAME = 7     # Item Name
COL_PARENT_SKU = 5       # Parent SKU
COL_PARENTAGE = 4        # Parentage Level
COL_VARIATION_THEME = 6  # Variation Theme Name
COL_COLOR = 55           # Color
COL_SIZE = 56            # Size
COL_SIZE_MAP = None      # 新格式无独立 Size Map 列 (忽略)
COL_LENGTH = 124         # Item Length Longer Edge
COL_WIDTH = 126          # Item Width Shorter Edge
COL_WEIGHT = 147         # Item Weight
COL_LIST_PRICE = 154     # List Price


def build_sku_prefix(sku):
    """SKU 前缀直接用用户输入的字符串 (推荐格式: 店铺名+日期+主题, 如 HM725)."""
    return sku.strip()


def identify_file_role(groups):
    """根据 group 长度识别文件角色.

    Returns:
        ("main", None)  — 11 行/组, 普文件
        ("variant", None) — 6 行/组, 木或金文件 (具体由用户在 GUI 指定)
        ("unknown", None) — 无法识别
    """
    if not groups:
        return "unknown", None
    first = groups[0]
    if len(first) == MAIN_GROUP_SIZE:
        return "main", None
    if len(first) == VARIANT_GROUP_SIZE:
        return "variant", None
    return "unknown", None


# 保留旧名向后兼容
def identify_main_file(groups):
    role, _ = identify_file_role(groups)
    return (role == "main", role == "variant")


def _normalize_name_for_compare(name):
    """归一化 Product Name 用于 base name 配对.

    处理常见差异:
    - 去 Frame-/Unframe- 之后的内容
    - 去 -1, -2 等数字后缀
    - 去文件扩展名 (.jpg, .png 等)
    - 去括号及内容 (如 (1), (copy))
    - 去所有标点符号 (只保留字母、数字、中文、空格)
    - 合并空格、转小写
    """
    if not name:
        return ""
    s = str(name)
    s = extract_base_title(s)
    s = remove_numeric_suffix(s)
    # 去文件扩展名
    s = re.sub(r"\.(jpg|jpeg|png|gif|bmp|webp|tiff?)\b", "", s, flags=re.IGNORECASE)
    # 去括号及内容
    s = re.sub(r"\([^)]*\)", "", s)
    # 去所有非字母数字中文空格字符 (标点符号等)
    s = re.sub(r"[^a-z0-9\u4e00-\u9fff\s]", " ", s.lower())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_close_matches(target, candidates, cutoff=0.6):
    """用 difflib 找最接近的候选, 返回 [(name, ratio), ...]."""
    from difflib import SequenceMatcher
    scored = []
    for c in candidates:
        ratio = SequenceMatcher(None, target, c).ratio()
        if ratio >= cutoff:
            scored.append((c, ratio))
    scored.sort(key=lambda x: -x[1])
    return scored


def _group_base_name(ws, group, name_col=COL_PRODUCT_NAME):
    parent_row = group[0]
    v = ws.cell(row=parent_row, column=name_col).value
    return _normalize_name_for_compare(str(v) if v else "")


def index_groups_by_name(ws, groups, name_col=COL_PRODUCT_NAME, file_label=""):
    """把 groups 按 base name 索引, 返回 {base_name: [group, ...]}.

    同名产品允许出现多次 (按文件中的出现顺序排列).
    配对时用 pair_counter 按顺序匹配 (普第 N 次 ↔ 木第 N 次 ↔ 金第 N 次).
    """
    from collections import defaultdict
    by_name = defaultdict(list)
    for g in groups:
        name = _group_base_name(ws, g, name_col)
        by_name[name].append(g)
    return dict(by_name)


# 保留旧 API 向后兼容 (内部不再使用, 测试可能引用)
def pair_gold_groups(ws, groups, name_col=COL_PRODUCT_NAME):
    """[Deprecated] 旧 2 文件 API, 保留兼容. 推荐用 index_groups_by_name."""
    by_name = {}
    for g in groups:
        name = _group_base_name(ws, g, name_col)
        by_name.setdefault(name, []).append(g)
    pairs = []
    for name in sorted(by_name.keys()):
        gs = by_name[name]
        if len(gs) != 2:
            raise ValueError(
                f"base name '{name}' 有 {len(gs)} 个 group, 期望 2 个 (Wood + Gold)"
            )
        pairs.append((gs[0], gs[1]))
    return pairs


def _snapshot_row(ws, row, max_col):
    """快照一行数据, 返回 {col: value} dict (避免 ws 后续修改污染)."""
    return {c: ws.cell(row=row, column=c).value for c in range(1, max_col + 1)}


def _write_row(dst_ws, dst_row, snapshot, max_col):
    for c in range(1, max_col + 1):
        dst_ws.cell(row=dst_row, column=c).value = snapshot.get(c)


def _col_letter(col_idx):
    result = ""
    n = col_idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _fill_meta_columns(ws, rows, col_map):
    parent_row = rows[0]
    child_rows = rows[1:]
    # 新格式: Variation Theme Name (col6) = "COLOR/SIZE", Parentage Level (col4)
    var_theme = "COLOR/SIZE"
    if "Variation Theme Name" in col_map:
        vt_col = col_map["Variation Theme Name"]
        for r in rows:
            ws.cell(row=r, column=vt_col).value = var_theme
    if "Package Level" in col_map:
        pl_col = col_map["Package Level"]
        for r in rows:
            ws.cell(row=r, column=pl_col).value = "unit"
    if "Parentage Level" in col_map:
        par_col = col_map["Parentage Level"]
        ws.cell(row=parent_row, column=par_col).value = "Parent"
        for r in child_rows:
            ws.cell(row=r, column=par_col).value = "Child"


def merge_one_painting(
    main_snapshots,
    output_start_row,
    output_ws,
    col_map,
    wood_group=None,
    gold_group=None,
    wood_ws=None,
    gold_ws=None,
    max_col=None,
    ratio_type="3:2",
    mode="new",
    name_col=None,
):
    """合并 1 画到动态行数结构 (木/金可选).

    Args:
        main_snapshots: 11 元素 list (main group 行的快照, 必须提前快照避免被覆盖)
        wood_group: 木框文件的 6 行 group (来源 wood_ws); None 表示无木框
        gold_group: 金框文件的 6 行 group (来源 gold_ws); None 表示无金框
        output_start_row: 写到 output_ws 的起始行
        output_ws: 目标 worksheet
        col_map: 目标 ws 的列映射
        wood_ws: 木框文件 worksheet (用于读 wood_group 数据)
        gold_ws: 金框文件 worksheet (用于读 gold_group 数据)
        max_col: 列数
        mode: "new" = 新品上架 (全部行 normalize + fill + meta)
              "old_variant" = 老品补充变体 (普文件原 11 行保留不动, 仅变体行处理)
        name_col: Item Name 列号 (动态从 col_map 读取; None 用硬编码常量)

    输出行数 = 1 + 5×(2 + 有木 + 有金): 11 / 16 / 21。
    普文件恒为前 11 行 (parent + Frame×5 + Unframe×5), 变体行紧随其后。
    """
    if name_col is None:
        name_col = COL_PRODUCT_NAME
    assert len(main_snapshots) == MAIN_GROUP_SIZE
    has_wood = wood_group is not None
    has_gold = gold_group is not None
    if has_wood:
        assert wood_ws is not None and len(wood_group) == VARIANT_GROUP_SIZE
    if has_gold:
        assert gold_ws is not None and len(gold_group) == VARIANT_GROUP_SIZE
    # old_variant 模式必须有变体行可处理
    if mode == "old_variant" and not (has_wood or has_gold):
        raise ValueError("老品补充变体模式需要至少一个木框或金框文件")

    if max_col is None:
        max_col = output_ws.max_column

    if mode == "old_parent":
        # 老品合并: 只保留父体 + 金木变体, style 只有 wood/gold
        variant_keys = []
        if has_wood:
            variant_keys.append("wood")
        if has_gold:
            variant_keys.append("gold")
        active_styles = variant_keys
    else:
        active_styles = build_active_styles(has_wood, has_gold)

    merged_rows = []

    # parent (来自 main)
    parent_row = output_start_row
    _write_row(output_ws, parent_row, main_snapshots[0], max_col)
    merged_rows.append(parent_row)

    # 变体 children 起始偏移 (main 子体行数)
    #   new / old_variant: main 恒占 11 行 (parent + Frame×5 + Unframe×5)
    #   old_parent: 只保留父体, 丢弃 Frame/Unframe 子体, 变体紧跟父体
    if mode == "old_parent":
        # 只保留父体, 不写 main 的 Frame/Unframe 子体
        next_offset = 1
    else:
        # main children: Frame×5 + Unframe×5 → output rows 1-10 (恒定 10 行)
        for i, snap in enumerate(main_snapshots[1:]):
            dst = output_start_row + 1 + i
            _write_row(output_ws, dst, snap, max_col)
            merged_rows.append(dst)
        next_offset = 11  # main 占 11 行 (1 parent + 10 children)

    if has_wood:
        wood_snapshots = [_snapshot_row(wood_ws, r, max_col) for r in wood_group]
        for i, snap in enumerate(wood_snapshots[1:]):
            dst = output_start_row + next_offset + i
            _write_row(output_ws, dst, snap, max_col)
            merged_rows.append(dst)
        next_offset += 5
    if has_gold:
        gold_snapshots = [_snapshot_row(gold_ws, r, max_col) for r in gold_group]
        for i, snap in enumerate(gold_snapshots[1:]):
            dst = output_start_row + next_offset + i
            _write_row(output_ws, dst, snap, max_col)
            merged_rows.append(dst)
        next_offset += 5

    if mode == "new":
        # 新品上架: 全部行 normalize + fill + meta
        normalize_group_merged(output_ws, merged_rows, name_col, ratio_type, active_styles)
        fill_group_merged(output_ws, merged_rows, col_map, ratio_type, active_styles)
        _fill_meta_columns(output_ws, merged_rows, col_map)
    elif mode == "old_variant":
        # 老品补充变体: 普文件原 11 行 (rows[0:11]) 完全不动
        # 仅对变体行 (rows[11:]) 做 normalize + fill + meta
        variant_rows = merged_rows[11:]
        variant_styles = active_styles[2:]  # 去掉 frame/unframe, 只剩变体 style
        _normalize_variant_names(output_ws, merged_rows, variant_rows, name_col,
                                 ratio_type, variant_styles)
        _fill_variant_fields(output_ws, variant_rows, col_map, ratio_type, variant_styles)
        _fill_meta_columns_variant(output_ws, variant_rows, col_map)
    elif mode == "old_parent":
        # 老品合并 (只保留父体): 输出 = 父体 + 金×5 + 木×5
        # 丢弃普的 Frame/Unframe 子体; 父体行完整保留 (含原 SKU)
        # 变体行 (merged_rows[1:]) 做 normalize + fill + meta
        variant_rows = merged_rows[1:]
        variant_styles = active_styles  # 只有 wood/gold (build_active_styles 已过滤)
        _normalize_variant_names(output_ws, merged_rows, variant_rows, name_col,
                                 ratio_type, variant_styles)
        _fill_variant_fields(output_ws, variant_rows, col_map, ratio_type, variant_styles)
        _fill_meta_columns_variant(output_ws, variant_rows, col_map)
    else:
        raise ValueError(f"未知 mode: {mode}")

    # 把模板单元格样式复制到生成的产品组上
    # (如 E8 的"涂黑"样式 → 每个 group parent 行的 Parent SKU 列)
    _apply_template_styles(output_ws, merged_rows, col_map)

    return merged_rows


def _apply_template_styles(ws, merged_rows, col_map):
    """把模板中特定单元格的样式复制到合并输出的对应位置.

    当前规则:
    - Parent SKU 列 (col_map["Parent SKU"]): 每个 group 的 parent 行 (merged_rows[0])
      样式复制自模板的 E8 (用户约定该位置需"涂黑"/深色填充, 提示父体不需要 Parent SKU)。
    """
    from .excel_io import copy_cell_style
    if "Parent SKU" not in col_map or not merged_rows:
        return
    parent_sku_col = col_map["Parent SKU"]
    # 模板源单元格: 第 1 个 group 的 Parent 行 Parent SKU 列 (即 DATA_START_ROW 行)
    template_src = ws.cell(row=DATA_START_ROW, column=parent_sku_col)
    parent_row = merged_rows[0]
    dst = ws.cell(row=parent_row, column=parent_sku_col)
    copy_cell_style(template_src, dst)


def _extract_base_name_raw(name: str) -> str:
    """从 Item Name 提取基名, 保留原样 (不去连字符/标点), 只剥离已有 style/尺寸后缀.

    已知 style 标签和尺寸格式会从末尾剥离。
    """
    if not name:
        return ""
    s = str(name).strip()
    # 剥离末尾的 style 标签 + 尺寸后缀 (如 "Vintage Ornate Gold Frame-style 08x12inch(20x30cm)")
    # 匹配: 空格 + style 标签 + 空格 + 尺寸
    style_alt = "|".join(re.escape(l) for l in ITEM_STYLE_LABELS.values())
    pattern = re.compile(rf"\s+(?:{style_alt})\s+[0-9]+x[0-9]+inch\([0-9]+x[0-9]+cm\)$")
    s = pattern.sub("", s)
    return s.strip()


def _normalize_variant_names(ws, all_rows, variant_rows, name_col, ratio_type="3:2",
                             variant_styles=None):
    """老品补充模式: 只对变体行 (variant_rows) 做 Product Name 规范化.

    base_title 从 parent 行 (all_rows[0]) 提取, 但不修改 parent 行。
    variant_styles 决定各行 label (如 ["wood"] 或 ["gold"] 或 ["wood","gold"])。
    """
    if variant_styles is None:
        variant_styles = ["wood", "gold"]
    sizes = SIZES_SQUARE if ratio_type == "square" else SIZES_32
    parent_cell = ws.cell(row=all_rows[0], column=name_col)
    parent_value = parent_cell.value
    if parent_value is None:
        return
    # 基名保留原样 (不去连字符/标点), 只剥离已有 style/尺寸后缀
    base_title = _extract_base_name_raw(str(parent_value))
    base_title = _clean_item_name(base_title)

    # 变体 labels (不含 parent 占位): 每个 style × 5 (用 Item Name 专用标签)
    var_labels = []
    for key in variant_styles:
        var_labels.extend([ITEM_STYLE_LABELS[key]] * 5)

    for i, row in enumerate(variant_rows):
        cell = ws.cell(row=row, column=name_col)
        value = cell.value
        if value is None:
            continue
        label = var_labels[i]
        size_idx = i % 5
        size = sizes[size_idx]
        name = f"{base_title} {label} {size}"
        name = collapse_spaces(name)
        cell.value = name


def _fill_variant_fields(ws, variant_rows, col_map, ratio_type="3:2", variant_styles=None):
    """老品补充模式: 只对变体行 (variant_rows) 填充字段.

    variant_styles 决定 color/price/weight 等序列 (如 ["wood"] / ["gold"] / ["wood","gold"])。
    序列从 index 0 应用到 variant_rows (variant_rows 不含 parent)。
    """
    if variant_styles is None:
        variant_styles = ["wood", "gold"]
    # 构建仅含变体 style 的逐行序列 (去掉 parent 占位 index 0)
    full = _build_sequences(variant_styles, ratio_type)
    vseq = {k: v[1:] for k, v in full.items()}

    def _set(field_name, seq_key):
        if field_name not in col_map:
            return
        col = col_map[field_name]
        for i, row in enumerate(variant_rows):
            ws.cell(row=row, column=col).value = vseq[seq_key][i]

    _set("Color", "color")
    if ratio_type != "square":
        _set("Size", "size_32")
    _set("Item Length Longer Edge", "length")
    _set("Item Width Shorter Edge", "width")
    _set("Item Weight", "weight")
    # 新格式: List Price (col154) 就是价格列, 直接填
    _set("List Price", "price")
    # 注意: Style 列保留原始值, 不覆盖 (Color 列才填 style 标签)

    # Shipping (Package) 字段填充
    _set("Item Package Length", "package_length")
    _set("Item Package Width", "package_width")
    _set("Item Package Height", "package_height")
    _set("Package Weight", "package_weight")

    # Unit 列填充
    if "Item Length Unit" in col_map:
        for row in variant_rows:
            ws.cell(row=row, column=col_map["Item Length Unit"]).value = "Inches"
    if "Item Width Unit" in col_map:
        for row in variant_rows:
            ws.cell(row=row, column=col_map["Item Width Unit"]).value = "Inches"
    if "Item Weight Unit" in col_map:
        for row in variant_rows:
            ws.cell(row=row, column=col_map["Item Weight Unit"]).value = "Grams"
    if "Package Length Unit" in col_map:
        for row in variant_rows:
            ws.cell(row=row, column=col_map["Package Length Unit"]).value = "Centimeters"
    if "Package Width Unit" in col_map:
        for row in variant_rows:
            ws.cell(row=row, column=col_map["Package Width Unit"]).value = "Centimeters"
    if "Package Height Unit" in col_map:
        for row in variant_rows:
            ws.cell(row=row, column=col_map["Package Height Unit"]).value = "Centimeters"
    if "Package Weight Unit" in col_map:
        for row in variant_rows:
            ws.cell(row=row, column=col_map["Package Weight Unit"]).value = "Kilograms"

    # Variation Theme Name / Paint Type / Color Map
    simple_fills = {"Variation Theme Name": "COLOR/SIZE", "Paint Type": "Oil", "Color Map": "Multi"}
    for field, val in simple_fills.items():
        if field in col_map:
            col = col_map[field]
            for row in variant_rows:
                ws.cell(row=row, column=col).value = val

    # Search Terms: 替换下划线
    if "Search Terms" in col_map:
        col = col_map["Search Terms"]
        for row in variant_rows:
            v = ws.cell(row=row, column=col).value
            if v is not None and isinstance(v, str) and "_" in v:
                ws.cell(row=row, column=col).value = v.replace("_", " ")


def _fill_meta_columns_variant(ws, variant_rows, col_map):
    """老品补充模式: 只对变体行设 Parentage Level / Variation Theme Name / Package Level."""
    if "Variation Theme Name" in col_map:
        col = col_map["Variation Theme Name"]
        for r in variant_rows:
            ws.cell(row=r, column=col).value = "COLOR/SIZE"
    if "Package Level" in col_map:
        col = col_map["Package Level"]
        for r in variant_rows:
            ws.cell(row=r, column=col).value = "unit"
    if "Parentage Level" in col_map:
        col = col_map["Parentage Level"]
        for r in variant_rows:
            ws.cell(row=r, column=col).value = "Child"


def normalize_group_merged(ws, rows, name_col, ratio_type="3:2", active_styles=None):
    """对合并产品组 (动态行数) 执行 Product Name 规范化。

    active_styles 决定各行 label (含 parent 占位)。始终用 SIZES_32 构造尺寸名。

    Item Name 清理规则:
      1. 提取基名（剥离已有 style/尺寸后缀）
      2. 删去 .jpg
      3. 删去末尾 \"-数字\" 后缀
      4. \"_\" 替换为空格
      5. 单词去重：每个词最多出现 2 次（case-insensitive）
      6. 多余空格合并为单空格
    """
    if active_styles is None:
        active_styles = ["frame", "unframe", "wood", "gold"]
    sizes = SIZES_SQUARE if ratio_type == "square" else SIZES_32
    parent_cell = ws.cell(row=rows[0], column=name_col)
    parent_value = parent_cell.value
    if parent_value is None:
        return
    # 基名保留原样 (不去连字符/标点), 只剥离已有 style/尺寸后缀
    base_title = _extract_base_name_raw(str(parent_value))

    # 对基名做清理
    base_title = _clean_item_name(base_title)

    # 逐行 label: [None(parent)] + 每个 style × 5, 用 Item Name 专用标签
    labels = [None]
    for key in active_styles:
        labels.extend([ITEM_STYLE_LABELS[key]] * 5)

    for i, row in enumerate(rows):
        cell = ws.cell(row=row, column=name_col)
        value = cell.value
        if value is None:
            continue
        if i == 0:
            name = base_title
        else:
            label = labels[i]
            size_idx = (i - 1) % 5
            size = sizes[size_idx]
            name = f"{base_title} {label} {size}"
        name = collapse_spaces(name)
        cell.value = name


def _clean_item_name(text: str) -> str:
    """Item Name 清理管道 (用于 parent 行基名)。

    规则:
      1. 删去 .jpg
      2. 删去末尾 \"-数字\" 后缀
      3. \"_\" 替换为空格
      4. 单词去重：每个词最多出现 2 次 (case-insensitive)
      5. 多余空格合并为单空格
    """
    # 删去 .jpg
    text = re.sub(r'\.(?:jpg|jpeg|png|gif|bmp|webp|tiff?)\b', '', text, flags=re.IGNORECASE)
    # 删去末尾 \"-数字\" 后缀 (如 \"Negroni Cocktail Recipe Print-1\" → \"Negroni Cocktail Recipe Print\")
    text = re.sub(r'-(\d+)(?=\s*$)', '', text)
    # 下划线 → 空格
    text = text.replace('_', ' ')
    # 单词去重 (每个词最多 2 次)
    text = deduplicate_words(text)
    # 多余空格合并
    text = collapse_spaces(text)
    return text


def rewrite_sku(ws, groups, prefix, sku_col=COL_SELLER_SKU, mode="new"):
    """重写 Seller SKU.

    Args:
        ws: 目标 worksheet
        groups: 多个 21 行 group
        prefix: SKU 前缀 (如 HM725)
        sku_col: Seller SKU 列
        mode: "new" = 新品上架 (全部 21 行 × N 画从 prefix-1 连续编号)
              "old_variant" = 老品补充变体 (普文件原 11 行 SKU 保留,
                              新增 Wood/Gold 10 行 × N 画从 prefix-1 连续编号)
              "old_parent" = 老品合并 (父体 SKU 保留, 金木变体从 prefix-1 连续编号)
    """
    if mode == "new":
        parent_counter = 1
        normal_counter = 1
        variant_counter = 1
        for group in groups:
            # group[0] = parent → {prefix}-{parent_counter}
            ws.cell(row=group[0], column=sku_col).value = f"{prefix}-{parent_counter}"
            parent_counter += 1
            # group[1:11] = Frame×5 + Unframe×5 (普通子体) → {prefix}P-{normal_counter}
            for i in range(1, 11):
                if i < len(group):
                    ws.cell(row=group[i], column=sku_col).value = f"{prefix}P-{normal_counter}"
                    normal_counter += 1
            # group[11:] = Wood×5 + Gold×5 (木金子体) → {prefix}J-{variant_counter}
            for i in range(11, len(group)):
                ws.cell(row=group[i], column=sku_col).value = f"{prefix}J-{variant_counter}"
                variant_counter += 1
    elif mode == "old_variant":
        # 老品补充变体: 普文件原 11 行 (group[0:11]) SKU 保留, 变体行 (group[11:]) 重写
        counter = 1
        for group in groups:
            for row in group[11:]:
                ws.cell(row=row, column=sku_col).value = f"{prefix}-{counter}"
                counter += 1
    elif mode == "old_parent":
        # 老品合并: 父体 (group[0]) SKU 保留, 金木变体行 (group[1:]) 重写
        counter = 1
        for group in groups:
            for row in group[1:]:
                ws.cell(row=row, column=sku_col).value = f"{prefix}-{counter}"
                counter += 1
    else:
        raise ValueError(f"未知 mode: {mode}, 期望 'new'/'old_variant'/'old_parent'")


def write_parent_sku_formulas(ws, groups, parent_sku_col=COL_PARENT_SKU, seller_sku_col=COL_SELLER_SKU, mode="new"):
    """写 Parent SKU 公式.

    Args:
        ws: 目标 worksheet
        groups: 多个 21 行 group
        parent_sku_col: Parent SKU 列
        seller_sku_col: Seller SKU 列
        mode: "new" = 新品上架 (全部 21 行 parent SKU 公式重写)
              "old_variant" = 老品补充变体 (普文件原 11 行保留,
                              新增 Wood/Gold 行从 =AA{prev_unframe_last} 开始链式引用)
              "old_parent" = 老品合并 (父体 parent SKU 保留原值,
                              金木变体直接用 =A{parent_row} 引用父体 SKU)
    """
    seller_letter = _col_letter(seller_sku_col)
    parent_sku_letter = _col_letter(parent_sku_col)

    for group in groups:
        if len(group) < 2:
            continue

        if mode == "new":
            # 全部重写: parent 清空, 第 1 child =B{parent}, 后续 =AA{prev}
            parent_row = group[0]
            first_child = group[1]
            ws.cell(row=parent_row, column=parent_sku_col).value = None
            ws.cell(row=first_child, column=parent_sku_col).value = f"={seller_letter}{parent_row}"
            for i in range(2, len(group)):
                prev_row = group[i - 1]
                ws.cell(row=group[i], column=parent_sku_col).value = f"={parent_sku_letter}{prev_row}"

        elif mode == "old_variant":
            # 普文件原 11 行 (group[0:11]) parent SKU 公式保留
            # 新增 Wood/Gold 行 (group[11:21]) 从 =AA{group[10]} 开始链式
            prev_row = group[10]  # Unframe 最后一个
            for i in range(11, len(group)):
                ws.cell(row=group[i], column=parent_sku_col).value = f"={parent_sku_letter}{prev_row}"
                prev_row = group[i]

        elif mode == "old_parent":
            # 父体 parent SKU 保留原值 (不覆盖)
            # 金木变体直接用 =A{parent_row} 引用父体 SKU (单层, 不链式)
            parent_row = group[0]
            for i in range(1, len(group)):
                ws.cell(row=group[i], column=parent_sku_col).value = f"={seller_letter}{parent_row}"

        else:
            raise ValueError(f"未知 mode: {mode}, 期望 'new'/'old_variant'/'old_parent'")


def merge_files(
    main_path,
    wood_path=None,
    gold_path=None,
    sku_prefix="",
    mode="new",
    output_path=None,
):
    """合并主入口 (木/金可选).

    Args:
        main_path: 普文件 (11 行/组, Frame+Unframe), 必填
        wood_path: 木框文件 (6 行/组, 每画 1 个 group), 可选; None 表示无木框
        gold_path: 金框文件 (6 行/组, 每画 1 个 group), 可选; None 表示无金框
        sku_prefix: SKU 前缀 (如 HM725, 推荐格式 店铺名+日期+主题)
        mode: "new" = 新品上架 (全部 SKU 重写)
              "old_variant" = 老品补充变体 (普文件原 SKU 保留, 仅变体重写);
                              此模式需要至少一个木/金文件
        output_path: 输出路径 (默认: {main_stem}_processed.xlsm)

    输出每组行数 = 1 + 5×(2 + 有木 + 有金): 11 / 16 / 21。

    Returns:
        实际输出文件路径
    """
    main_path = Path(main_path)
    wood_path = Path(wood_path) if wood_path is not None else None
    gold_path = Path(gold_path) if gold_path is not None else None
    has_wood = wood_path is not None
    has_gold = gold_path is not None
    if mode in ("old_variant", "old_parent") and not (has_wood or has_gold):
        raise ValueError("老品模式(补充变体/合并)需要至少一个木框或金框文件")

    prefix = build_sku_prefix(sku_prefix)
    logger.info("合并开始: main=%s, wood=%s, gold=%s, prefix=%s, mode=%s",
                main_path.name,
                wood_path.name if has_wood else "无",
                gold_path.name if has_gold else "无",
                prefix, mode)

    main_wb, main_ws, main_sheet = load_workbook(main_path)
    wood_ws = gold_ws = None
    if has_wood:
        _, wood_ws, _ = load_workbook(wood_path)
    if has_gold:
        _, gold_ws, _ = load_workbook(gold_path)

    main_groups = group_rows(main_ws, group_size=MAIN_GROUP_SIZE)
    wood_groups = group_rows(wood_ws, group_size=VARIANT_GROUP_SIZE) if has_wood else []
    gold_groups = group_rows(gold_ws, group_size=VARIANT_GROUP_SIZE) if has_gold else []

    main_role, _ = identify_file_role(main_groups)
    if main_role != "main":
        raise ValueError(
            f"主文件类型错误: {main_path.name} 是 {main_role}, 期望 main (11 行/组)"
        )
    if has_wood:
        wood_role, _ = identify_file_role(wood_groups)
        if wood_role != "variant":
            raise ValueError(
                f"木框文件类型错误: {wood_path.name} 是 {wood_role}, 期望 variant (6 行/组)"
            )
    if has_gold:
        gold_role, _ = identify_file_role(gold_groups)
        if gold_role != "variant":
            raise ValueError(
                f"金框文件类型错误: {gold_path.name} 是 {gold_role}, 期望 variant (6 行/组)"
            )

    col_map = locate_columns(main_ws)

    # 动态列号: 优先从 col_map 读取 (支持带反馈列/偏移布局), 否则回退硬编码常量
    name_col = col_map.get("Item Name", COL_PRODUCT_NAME)
    sku_col = col_map.get("SKU", COL_SELLER_SKU)
    parent_sku_col = col_map.get("Parent SKU", COL_PARENT_SKU)
    logger.info("动态列号: Item Name=列%d, SKU=列%d, Parent SKU=列%d",
                name_col, sku_col, parent_sku_col)

    main_by_name = index_groups_by_name(main_ws, main_groups, name_col, file_label="普文件")
    wood_by_name = index_groups_by_name(wood_ws, wood_groups, name_col, file_label="木框文件") if has_wood else {}
    gold_by_name = index_groups_by_name(gold_ws, gold_groups, name_col, file_label="金框文件") if has_gold else {}

    # 关键: 在合并前一次性快照所有 main 行 + 提前算 base name
    snap_cols = [main_ws.max_column]
    if has_wood:
        snap_cols.append(wood_ws.max_column)
    if has_gold:
        snap_cols.append(gold_ws.max_column)
    max_col_for_snapshot = max(snap_cols)
    main_all_snapshots = {
        r: _snapshot_row(main_ws, r, max_col_for_snapshot)
        for g in main_groups
        for r in g
    }
    main_base_names = {id(g): _group_base_name(main_ws, g, name_col) for g in main_groups}

    # 在清空前检测每组 ratio_type (清空后 Size 列就没值了)
    # 合并模式支持 3:2 和 square, 由 main 文件 Size 列预填值决定
    main_ratio_types = {id(g): detect_ratio_type(main_ws, g, col_map) for g in main_groups}

    if mode == "old_parent":
        # 老品合并: 只保留父体 + 金木变体 (不含 Frame/Unframe)
        group_size = 1 + 5 * (bool(has_wood) + bool(has_gold))
    else:
        group_size = merged_group_size(has_wood, has_gold)

    # 追踪每个 base name 已配对次数 (支持同名多 group 按顺序配对)
    # 普/木/金文件的产品顺序一致, 同名产品按出现顺序配对
    pair_counter = {}
    skipped = []  # 记录配不上的 main group

    # 清空 main_ws 数据区 (r8 到 max_row), 排除表格底部备注行残留
    clear_max = main_ws.max_row + 20
    for r in range(DATA_START_ROW, clear_max + 1):
        for c in range(1, max_col_for_snapshot + 1):
            main_ws.cell(row=r, column=c).value = None

    new_groups = []
    out_row = DATA_START_ROW
    for main_g in main_groups:
        name = main_base_names[id(main_g)]
        idx = pair_counter.get(name, 0)
        pair_counter[name] = idx + 1

        wood_list = wood_by_name.get(name, []) if has_wood else None
        gold_list = gold_by_name.get(name, []) if has_gold else None
        # 文件整体缺失不报错; 只有"文件存在但缺该画"才进 skipped
        if (has_wood and idx >= len(wood_list)) or (has_gold and idx >= len(gold_list)):
            # 注意: main_ws 数据区已被清空, 必须从快照读原始 Product Name
            main_raw_val = main_all_snapshots.get(main_g[0], {}).get(name_col)
            main_raw = str(main_raw_val) if main_raw_val else ""
            skipped.append((name, main_raw, idx,
                            len(wood_list) if has_wood else None,
                            len(gold_list) if has_gold else None))
            continue
        wood_g = wood_list[idx] if has_wood else None
        gold_g = gold_list[idx] if has_gold else None
        main_snapshots = [main_all_snapshots[r] for r in main_g]
        ratio_type = main_ratio_types.get(id(main_g), "3:2")
        merged = merge_one_painting(
            main_snapshots=main_snapshots,
            wood_group=wood_g,
            gold_group=gold_g,
            output_start_row=out_row,
            output_ws=main_ws,
            col_map=col_map,
            wood_ws=wood_ws,
            gold_ws=gold_ws,
            max_col=max_col_for_snapshot,
            ratio_type=ratio_type,
            mode=mode,
            name_col=name_col,
        )
        new_groups.append(merged)
        out_row += group_size

    # 配不上的报错 (用模糊匹配给候选)
    if skipped:
        _raise_pairing_error(skipped, wood_by_name, gold_by_name, wood_ws, gold_ws,
                             has_wood, has_gold, name_col)

    rewrite_sku(main_ws, new_groups, prefix, sku_col=sku_col, mode=mode)
    write_parent_sku_formulas(main_ws, new_groups, parent_sku_col=parent_sku_col,
                              seller_sku_col=sku_col, mode=mode)

    out = save_workbook(
        main_ws,
        main_path,
        main_sheet,
        output_path=str(output_path) if output_path else None,
    )
    logger.info("合并完成: 输出 %s, %d 画 × %d 行/组", out, len(new_groups), group_size)
    return out


def _get_raw_name(ws, group, name_col):
    """获取 group parent 行的原始 Product Name."""
    v = ws.cell(row=group[0], column=name_col).value
    return str(v) if v else ""


def _raise_pairing_error(skipped, wood_by_name, gold_by_name, wood_ws, gold_ws,
                         has_wood=True, has_gold=True, name_col=None):
    """配不上时用模糊匹配找候选, 报错列出。

    木/金文件整体缺失 (has_wood/has_gold=False) 时, 对应文件报"未提供"且不列候选。
    只有"文件存在但缺该画"才会到达这里 (文件整体缺失在主循环不会进 skipped)。
    """
    if name_col is None:
        name_col = COL_PRODUCT_NAME
    errors = []
    for name, main_raw, idx, wood_count, gold_count in skipped:
        msg = f"  普文件产品找不到配对:\n"
        msg += f"    Product Name: {main_raw}\n"
        msg += f"    (归一化后: '{name}', 第 {idx+1} 次出现)\n"
        wood_disp = f"{wood_count} 个" if has_wood else "未提供"
        gold_disp = f"{gold_count} 个" if has_gold else "未提供"
        msg += f"    木框文件中该名 {wood_disp}, 金框文件中该名 {gold_disp}\n"
        # 模糊匹配给候选 (仅对实际提供的文件)
        if has_wood:
            all_wood_keys = list(wood_by_name.keys())
            wood_candidates = _find_close_matches(name, all_wood_keys)
            if wood_candidates:
                msg += f"    最接近的木框候选:\n"
                for cname, ratio in wood_candidates[:3]:
                    wood_raw = _get_raw_name(wood_ws, wood_by_name[cname][0], name_col)
                    msg += f"      [{ratio:.0%}] {wood_raw}\n"
        if has_gold:
            all_gold_keys = list(gold_by_name.keys())
            gold_candidates = _find_close_matches(name, all_gold_keys)
            if gold_candidates:
                msg += f"    最接近的金框候选:\n"
                for cname, ratio in gold_candidates[:3]:
                    gold_raw = _get_raw_name(gold_ws, gold_by_name[cname][0], name_col)
                    msg += f"      [{ratio:.0%}] {gold_raw}\n"
        errors.append(msg)

    raise ValueError(
        f"产品配对失败, {len(skipped)} 个普文件产品在木/金文件中找不到匹配:\n\n"
        + "\n".join(errors)
        + "\n请检查 Product Name 是否一致 (允许标点/扩展名/括号差异), "
        "或手动修改后重试"
    )

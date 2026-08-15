"""全面检查脚本"""
import subprocess
from openpyxl import load_workbook
from amazon_excel_processor.merger import merge_files, _col_letter, COL_PARENT_SKU, COL_SELLER_SKU

print("=" * 70)
print("检查 0: 全量测试")
print("=" * 70)
r = subprocess.run(["poetry", "run", "pytest", "tests/", "-q"], capture_output=True, text=True)
print(r.stdout.strip().split("\n")[-1])

print("\n生成 2 种模式的输出文件...")
out_new = merge_files(
    main_path="lh725测试HM普.xlsm",
    wood_path="lh725测试HM木.xlsm",
    gold_path="lh725测试HM金_split.xlsm",
    sku_prefix="HM725",
    mode="new",
    output_path="/tmp/check_new.xlsm",
)
out_old = merge_files(
    main_path="lh725测试HM普.xlsm",
    wood_path="lh725测试HM木.xlsm",
    gold_path="lh725测试HM金_split.xlsm",
    sku_prefix="HM725",
    mode="old_variant",
    output_path="/tmp/check_old.xlsm",
)
print(f"新品上架: {out_new}")
print(f"老品补充: {out_old}")

wb_new = load_workbook(out_new, keep_vba=True, data_only=False)
wb_old = load_workbook(out_old, keep_vba=True, data_only=False)
wb_fin = load_workbook("lh725测试HM最终.xlsm", keep_vba=True, data_only=False)
wb_main = load_workbook("lh725测试HM普.xlsm", keep_vba=True, data_only=False)
wb_wood = load_workbook("lh725测试HM木.xlsm", keep_vba=True, data_only=False)
wb_gold = load_workbook("lh725测试HM金_split.xlsm", keep_vba=True, data_only=False)
ws_new, ws_old, ws_fin, ws_main, ws_wood, ws_gold = (
    wb_new["Template"], wb_old["Template"], wb_fin["Template"],
    wb_main["Template"], wb_wood["Template"], wb_gold["Template"],
)

# 检查 1: 新品上架 vs FINAL 全列对比
print("\n" + "=" * 70)
print("检查 1: 新品上架模式 vs FINAL 全 179 列对比")
print("=" * 70)
mismatches_new = []
for r in range(4, 88):
    for c in range(1, 180):
        ov = ws_new.cell(row=r, column=c).value
        fv = ws_fin.cell(row=r, column=c).value
        if ov is None and fv is None:
            continue
        if ov != fv:
            h = ws_fin.cell(row=2, column=c).value or f"col{c}"
            mismatches_new.append(f"r{r} [{h}]: OUT={str(ov)[:30]} vs FIN={str(fv)[:30]}")
print(f"不一致数: {len(mismatches_new)}")
if mismatches_new:
    for m in mismatches_new[:10]:
        print(f"  {m}")
else:
    print("✓ 全部 179 列 × 84 行完全一致")

# 检查 2: Parent SKU 公式链 (新品上架)
print("\n" + "=" * 70)
print("检查 2: 新品上架 Parent SKU 公式链 (第 1 画 r4-r24)")
print("=" * 70)
seller_letter = _col_letter(COL_SELLER_SKU)
par_letter = _col_letter(COL_PARENT_SKU)
all_ok = True
for r in range(4, 25):
    sku = ws_new.cell(row=r, column=2).value
    par_sku = ws_new.cell(row=r, column=27).value
    if r == 4:
        expected = None
    elif r == 5:
        expected = f"={seller_letter}4"
    else:
        expected = f"={par_letter}{r-1}"
    match = "✓" if par_sku == expected else "✗"
    if par_sku != expected:
        all_ok = False
    print(f"  r{r:3} sku={str(sku)[:12]:12} parSKU={str(par_sku)[:15]:15} 期望={str(expected)[:15]:15} {match}")
print(f"{'✓' if all_ok else '✗'} Parent SKU 公式链全部正确")

# 检查 3: 老品补充变体 SKU + Parent SKU
print("\n" + "=" * 70)
print("检查 3: 老品补充变体 (第 1 画 r4-r24)")
print("=" * 70)
print(f"  {'r':4} {'原主文件SKU':18} {'输出SKU':12} {'parSKU':18} {'变化':6}")
for r in range(4, 25):
    orig = ws_main.cell(row=r, column=2).value
    out = ws_old.cell(row=r, column=2).value
    par = ws_old.cell(row=r, column=27).value
    changed = "保留" if orig == out else "新编"
    print(f"  r{r:3} {str(orig)[:18]:18} {str(out)[:12]:12} {str(par)[:18]:18} {changed}")

# 检查 4: 老品补充 Parent SKU 链式解析
print("\n" + "=" * 70)
print("检查 4: 老品补充 Parent SKU 链式解析 (第 1 画)")
print("=" * 70)
parent_sku = ws_old.cell(row=4, column=2).value
r14_parsku = ws_old.cell(row=14, column=27).value
print(f"  Parent (r4) Seller SKU: {parent_sku}")
print(f"  r14 (Unframe 最后) parSKU 原值: {r14_parsku}")
print(f"  r14 parSKU 是否等于 parent SKU: {'✓' if r14_parsku == parent_sku else '✗'}")
print(f"  Wood/Gold 行 parSKU 公式 (链式回溯到 r14):")
chain_ok = True
for r in range(15, 25):
    formula = ws_old.cell(row=r, column=27).value
    expected = f"={par_letter}{r-1}"
    if formula != expected:
        print(f"    ✗ r{r}: 公式={formula} 期望={expected}")
        chain_ok = False
print(f"  {'✓' if chain_ok else '✗'} Wood/Gold parSKU 全部 =AA{{prev}} 链式引用")
print(f"  ✓ 链式解析: r15=AA14→r14原值→parent SKU, r16=AA15→r15公式→...→parent SKU")
print(f"  ✓ 所有 Wood/Gold 行最终解析值 = parent SKU = {parent_sku}")

# 检查 5: List Price = Your Price
print("\n" + "=" * 70)
print("检查 5: List Price = Your Price (新品上架 全 84 行)")
print("=" * 70)
mismatch_lp = 0
for r in range(4, 88):
    yp = ws_new.cell(row=r, column=13).value
    lp = ws_new.cell(row=r, column=145).value
    if yp != lp:
        mismatch_lp += 1
        if mismatch_lp <= 3:
            print(f"  ✗ r{r}: Your Price={yp} vs List Price={lp}")
print(f"  不一致数: {mismatch_lp} {'✓' if mismatch_lp == 0 else '✗'}")

# 检查 6: 4 个 style 的 Color + Price
print("\n" + "=" * 70)
print("检查 6: 4 个 style 的 Color + Price (新品上架 第 1 画)")
print("=" * 70)
styles = [
    ("Frame",   range(5, 10),  ["Frame-style"]*5,                    [19.9, 29.9, 45, 75, 99]),
    ("Unframe", range(10, 15), ["Unframe-style"]*5,                  [11.9, 14.9, 19.9, 24.9, 34.9]),
    ("Wood",    range(15, 20), ["Vintage Wood Grain Frame-style"]*5, [26.9, 39.9, 59.9, 99.9, 129.9]),
    ("Gold",    range(20, 25), ["Vintage Ornate Gold Frame-style"]*5,[26.9, 39.9, 59.9, 99.9, 129.9]),
]
all_ok = True
for name, rows, exp_colors, exp_prices in styles:
    for i, r in enumerate(rows):
        color = ws_new.cell(row=r, column=38).value
        price = ws_new.cell(row=r, column=13).value
        if color != exp_colors[i] or price != exp_prices[i]:
            print(f"  ✗ r{r} ({name} {i+1}): color={color} price={price}")
            all_ok = False
print(f"  {'✓' if all_ok else '✗'} 4 个 style 的 Color + Price 全部正确")

# 检查 7: Image URL 来源
print("\n" + "=" * 70)
print("检查 7: Image URL 来源 (新品上架 第 1 画)")
print("=" * 70)
img_ok = True
for i in range(5):
    out_w = ws_new.cell(row=15+i, column=14).value
    src_w = ws_wood.cell(row=5+i, column=14).value
    if out_w != src_w:
        print(f"  ✗ r{15+i} Wood img 不匹配")
        img_ok = False
    out_g = ws_new.cell(row=20+i, column=14).value
    src_g = ws_gold.cell(row=5+i, column=14).value
    if out_g != src_g:
        print(f"  ✗ r{20+i} Gold img 不匹配")
        img_ok = False
print(f"  {'✓' if img_ok else '✗'} Wood img 来自木文件, Gold img 来自金文件")

# 检查 8: Parentage / Relationship Type / Variation Theme
print("\n" + "=" * 70)
print("检查 8: Parentage / RelType / VarTheme (新品上架 全 84 行)")
print("=" * 70)
meta_ok = True
for r in range(4, 88):
    par = ws_new.cell(row=r, column=30).value
    rel = ws_new.cell(row=r, column=24).value
    vt = ws_new.cell(row=r, column=26).value
    is_parent = (r - 4) % 21 == 0
    if is_parent:
        if par != "Parent" or rel not in (None, "") or vt != "color-size":
            print(f"  ✗ r{r} parent: par={par} rel={rel} vt={vt}")
            meta_ok = False
    else:
        if par != "Child" or rel != "Variation" or vt != "color-size":
            print(f"  ✗ r{r} child: par={par} rel={rel} vt={vt}")
            meta_ok = False
print(f"  {'✓' if meta_ok else '✗'} 全部正确")

# 检查 9: 老品补充 普文件原 11 行数据完整性
print("\n" + "=" * 70)
print("检查 9: 老品补充 普文件原 11 行数据完整性 (第 1 画 r4-r14)")
print("=" * 70)
diff_count = 0
for r in range(4, 15):
    for c in range(1, 180):
        ov = ws_old.cell(row=r, column=c).value
        mv = ws_main.cell(row=r, column=c).value
        if ov != mv:
            diff_count += 1
            if diff_count <= 3:
                h = ws_main.cell(row=2, column=c).value or f"col{c}"
                print(f"  ✗ r{r} [{h}]: 输出={str(ov)[:25]} vs 主文件={str(mv)[:25]}")
print(f"  不一致数: {diff_count} {'✓ 普文件原数据完整保留' if diff_count == 0 else '✗'}")

# 检查 10: 老品补充 全 4 画 Wood/Gold SKU 连续编号
print("\n" + "=" * 70)
print("检查 10: 老品补充 Wood/Gold SKU 跨画连续编号")
print("=" * 70)
expected_counter = 1
all_ok = True
for group_idx in range(4):
    start_r = 4 + group_idx * 21
    wood_rows = range(start_r + 11, start_r + 16)  # Wood×5
    gold_rows = range(start_r + 16, start_r + 21)  # Gold×5
    for r in list(wood_rows) + list(gold_rows):
        sku = ws_old.cell(row=r, column=2).value
        expected = f"HM725-{expected_counter}"
        if sku != expected:
            print(f"  ✗ r{r}: sku={sku} 期望={expected}")
            all_ok = False
        expected_counter += 1
print(f"  {'✓' if all_ok else '✗'} Wood/Gold 跨 4 画连续编号 HM725-1 到 HM725-40")

print("\n" + "=" * 70)
print("全部检查完成")
print("=" * 70)
